# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 12:31:23 2019

@author: 212566876
"""

#import seaborn as sns
import matplotlib.pyplot as plt
from matplotlib.ticker import FuncFormatter
import openpyxl
import pandas as pd
import numpy as np

SMALL_SIZE = 8
MEDIUM_SIZE = 10
BIGGER_SIZE = 12

plt.rc('font', size=SMALL_SIZE)          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=SMALL_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=SMALL_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title

# load workbook
# wb = openpyxl.load_workbook('LAR country data and plots.xlsx', data_only=True)
# ws = wb['Carbon Tax Lifting Cost Impact']

# load workbook
wb = openpyxl.load_workbook('./D18 OPGEE.xlsm', data_only=True)
ws = wb['Results']

# read in needed columns for plto data
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
GOR = []            # gas oil ratio (scf/bbl)
WOR = []            # water oil ratio (bbl water/bbl oil)
WIR = []            # water injection ratio (bbl water/bbl oil)
api_gravity = []            # api gravity of crude
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []
flaring_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    api_gravity.append(ws.cell(row=34, column=col).value)
    GOR.append(ws.cell(row=46, column=col).value)
    WOR.append(ws.cell(row=47, column=col).value)
    WIR.append(ws.cell(row=48, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)
    flaring_CI.append(ws.cell(row=180, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1      
        

# convert lists to pandas dataframe
D18_data = {'Country':countries[:7], 'Field':fields[:7], 'Production':production[:7], \
        'GOR':GOR[:7], 'WOR':WOR[:7], 'WIR':WIR[:7], \
        'API Gravity':api_gravity[:7], 'Flaring (scf/bbl)':flaring_ratio[:7], \
        'Flaring CI':flaring_CI[:7], \
        'Exploration Combustion CI':exp_comb[:7], \
        'Exploration VFF CI':exp_vff[:7], \
        'Drilling Combustion CI':dnd_comb[:7], \
        'Drilling VFF CI':dnd_vff[:7], \
        'Production Combustion CI':prod_comb[:7], \
        'Production VFF CI':prod_vff[:7], \
        'Processing Combustion CI':proc_comb[:7], \
        'Processing VFF CI':proc_vff[:7], \
        'Maintenance Combustion CI':main_comb[:7], \
        'Maintenance VFF CI':main_vff[:7], \
        'Transportation CI':trans[:7], \
        'Miscellaneous CI':misc[:7], \
        'Offsite CI':offsite[:7], \
        'Total CI':net_CI[:7]
    }


D18_emissions = pd.DataFrame(D18_data)
D18_emissions['WIR'] = pd.to_numeric(D18_emissions['WIR'], errors='coerce')
D18_emissions['Total VFF CI'] = D18_emissions['Exploration VFF CI'] + \
                            D18_emissions['Drilling VFF CI'] + \
                            D18_emissions['Production VFF CI'] + \
                            D18_emissions['Processing VFF CI'] + \
                            D18_emissions['Maintenance VFF CI']

#################################################################
# convert gCO2e/MJ to tCO2e/bbl (for use if necessary)
#################################################################

# get fuel spec table imported into dataframe
fsSheet = wb['Fuel Specs']
api = np.zeros((42,))
lhv = np.zeros((42,))
for i in range(api.shape[0]):
    api[i] = fsSheet.cell(row=i+24, column=2).value
    lhv[i] = fsSheet.cell(row=i+24, column=12).value

# convert CIs into tCO2e/bbl for a copy of dataframe in te/bbl units
D18_emissions['LHV'] = np.interp(D18_emissions['API Gravity'], api, lhv)
D18_emissions_bbl = D18_emissions.copy()
parameter_lst = list(D18_emissions.columns)[9:-1]
for param in parameter_lst:
    D18_emissions_bbl[param] = D18_emissions[param] * D18_emissions['LHV'] / 1e6

# back convert te/bbl to g/MJ for flaring CI that's originally in te/bbl
D18_emissions['Flaring CI'] = D18_emissions['Flaring CI'] * 1e6 / D18_emissions['LHV']

# Add a column that shows methane CI (venting and fugitives) in both dataframes
D18_emissions['Venting & Fugitives CI'] = D18_emissions['Total VFF CI'] - D18_emissions['Flaring CI']
D18_emissions_bbl['Venting & Fugitives CI'] = D18_emissions_bbl['Total VFF CI'] - D18_emissions_bbl['Flaring CI']

# SANGEA average CI
SANGEA_CI_mmscf = 85    # tCO2e/mmscf gas produced
GOR_2017 = 3049         # scf/bbl
SANGEA_CI = SANGEA_CI_mmscf * GOR_2017 / D18_emissions['LHV'].mean()    # convert to g/MJ

# Sort values by year
D18_emissions = D18_emissions.sort_values(by='Field')

#########################################################
# generate emissions breakdown plot for each year (g/MJ)
#########################################################
N = len(D18_emissions.index)
ind = np.arange(N)
prod = D18_emissions['Production']
gor = D18_emissions['GOR']
flr = D18_emissions['Flaring (scf/bbl)']
wp = D18_emissions['WOR']*prod
wi = D18_emissions['WIR']*prod
expCI = D18_emissions['Exploration Combustion CI']
dndCI = D18_emissions['Drilling Combustion CI']
prodCI = D18_emissions['Production Combustion CI']
procCI = D18_emissions['Processing Combustion CI']
mainCI = D18_emissions['Maintenance Combustion CI']
vffCI = D18_emissions['Total VFF CI']
fCI = D18_emissions['Flaring CI']
vfCI = D18_emissions['Venting & Fugitives CI']
miscCI = D18_emissions['Miscellaneous CI']
transCI = D18_emissions['Transportation CI']
offsiteCI = D18_emissions['Offsite CI']
totalCI = D18_emissions['Total CI']

yearLabels = D18_emissions['Field'].tolist()

fig, (ax, ax2) = plt.subplots(2, 1, sharex=True)

# SUBPLOT 1
width = 0.75
#exp_bar = ax.bar(ind, expCI, width, label='Exploration', \
#                 color=(45/255, 96/255, 179/255, 0.8))
#dnd_bar = ax.bar(ind, dndCI, width, label='Drilling', \
#                 color=(232/255, 159/255, 12/255, 0.8))
prod_bar = ax.bar(ind, prodCI, width, label='Production', \
                  color=(7/255, 71/255, 77/255, 0.8))
proc_bar = ax.bar(ind, procCI, width, bottom=prodCI, \
                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
f_bar = ax.bar(ind, fCI, width, bottom=prodCI + procCI, \
            label='Flaring', color=(255/255, 0, 0, 0.8))
vf_bar = ax.bar(ind, vfCI, width, bottom=prodCI + procCI + fCI, \
            label='Venting & Fugitives', color=(125/255, 0, 0, 0.8))
misc_bar = ax.bar(ind, miscCI, width, bottom=prodCI + procCI + fCI + \
            vfCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
#trans_bar = ax.bar(ind, transCI, width, bottom=dndCI + prodCI + procCI + \
#            mainCI + vffCI + miscCI, label='Transport', \
#            color=(69/255,74/255,65/255,0.8))
#offsite_bar = ax.bar(ind, offsiteCI, width, bottom=prodCI + procCI + \
#            fCI + vfCI + miscCI, label='Offsite', \
#            color=(0,128/255,1,0.8))
#totalCI_line = ax.plot(ind, totalCI, 'kX-', label='Total CI')
# around the world comparison
sangea_line = ax.hlines(SANGEA_CI, ind[0], ind[-1], colors='b', linestyles='dotted', label='SANGEA CI')


ax.set_title('D18 Carbon Intensity Profile 2013-19')
ax.set_xticks(ind.tolist())
ax.set_xticklabels(yearLabels)
ax.set_ylabel('Carbon Intensity\n(gCO2e/MJ crude oil)')

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.7, box.height])

handles, labels = ax.get_legend_handles_labels()
handles = handles[1:] + [handles[0]]
labels = labels[1:] + [labels[0]]
ax.legend(handles[::-1], labels[::-1], loc='center left', \
          bbox_to_anchor=(1, 0.5))

# SUBPLOT 2
oil_line = ax2.plot(ind, prod, 'go-', label='Oil Prod. (bbl/d)')
gor_line = ax2.plot(ind, gor, 'o-', color=(125/255,0,0,0.8), label='GOR (scf/bbl)')
flr_line = ax2.plot(ind, flr, 'ro-', label='Flaring (scf/bbl)')
wp_line = ax2.plot(ind, wp, 'bo-', label='Water Prod. (bbl/d)')
ax2s = ax2.twinx()      # secondary axis for water inj
wi_line = ax2s.plot(ind, wi, 'o-', color=(0,0,125/255,0.8), label='Water Inj. (bbl/d)')

ax2.set_title('D18 Field Profile 2013-19')
ax2.set_xticks(ind.tolist())
ax2.set_xticklabels(yearLabels)
ax2.set_ylabel('Oil Prod., GOR\nFlaring, Water Prod.')
ax2.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))
#ax2.set_ylabel('Oil Production (bbl/d)')
ax2s.set_ylabel('Water Inj. (bbl/d)')
ax2s.yaxis.set_major_formatter(FuncFormatter(lambda x, p: format(int(x), ',')))

# compile legends
lns = oil_line+gor_line+flr_line+wp_line+wi_line
labs = [l.get_label() for l in lns]

# Shrink current axis by 20%
box = ax2.get_position()
ax2.set_position([box.x0, box.y0, box.width * 0.7, box.height])
ax2.legend(lns, labs, loc='center left', bbox_to_anchor=(1.2, 0.5))

fig.savefig('D18_field_and_emissions_profile_2013_to_2019.png', bbox_inches='tight', dpi=600)

#########################################
# get country averages and do comparison
#########################################

# convert lists to pandas dataframe
emissions_data = {'Country':['D18 2019']+countries[7:], 'Field':fields[6:], 'Production':production[6:], \
        'GOR':GOR[6:], 'WOR':WOR[6:], 'WIR':WIR[6:], \
        'API Gravity':api_gravity[6:], 'Flaring (scf/bbl)':flaring_ratio[6:], \
        'Flaring CI':flaring_CI[6:], \
        'Exploration Combustion CI':exp_comb[6:], \
        'Exploration VFF CI':exp_vff[6:], \
        'Drilling Combustion CI':dnd_comb[6:], \
        'Drilling VFF CI':dnd_vff[6:], \
        'Production Combustion CI':prod_comb[6:], \
        'Production VFF CI':prod_vff[6:], \
        'Processing Combustion CI':proc_comb[6:], \
        'Processing VFF CI':proc_vff[6:], \
        'Maintenance Combustion CI':main_comb[6:], \
        'Maintenance VFF CI':main_vff[6:], \
        'Transportation CI':trans[6:], \
        'Miscellaneous CI':misc[6:], \
        'Offsite CI':offsite[6:], \
        'Total CI':net_CI[6:]
    }

emissions = pd.DataFrame(emissions_data)
emissions['WIR'] = pd.to_numeric(emissions['WIR'], errors='coerce')
emissions['Total VFF CI'] = emissions['Exploration VFF CI'] + \
                            emissions['Drilling VFF CI'] + \
                            emissions['Production VFF CI'] + \
                            emissions['Processing VFF CI'] + \
                            emissions['Maintenance VFF CI']

emissions['LHV'] = np.interp(emissions['API Gravity'], api, lhv)
emissions['Flaring CI'] = emissions['Flaring CI'] * 1e6 / emissions['LHV']
emissions['Venting & Fugitives CI'] = emissions['Total VFF CI'] - emissions['Flaring CI']


# get country averages
def wavg(group, avg_name, weight_name):
    """ http://stackoverflow.com/questions/10951341/pandas-dataframe-aggregate-function-using-multiple-columns
    In rare instance, we may not have weights, so just return the mean. Customize this if your business case
    should return otherwise.
    """
    d = group[avg_name]
    w = group[weight_name]
    try:
        return (d * w).sum() / w.sum()
    except ZeroDivisionError:
        return d.mean()

parameter_lst2 = list(emissions.columns)[3:]
country_data = {}
country_data['Production'] = emissions.groupby('Country')['Production'].sum()
for param in parameter_lst2:
    country_data[param] = emissions.groupby('Country').apply(wavg, param, 'Production')

country_emissions = pd.DataFrame(data=country_data)     # country averages (g/MJ units)

# Add Venezuela data from Hassan's evaluation with public and commerical datasets
venezuela = pd.Series({'Drilling Combustion CI': 1.550958147, \
                       'Production Combustion CI': 6.749531274, \
                       'Processing Combustion CI': 2.389238867, \
                       'Total VFF CI': 5.6783, \
                       'Miscellaneous CI': 0.5, \
                       'Offsite CI': 1.845538113, \
                       'Total CI': 18.714, \
                       'API Gravity': 17.171, \
                       'GOR': 810.33, \
                       'WOR': 12.152, \
                       'Flaring (scf/bbl)': 227.94}, name='Venezuela')
venezuela['LHV'] = np.interp(venezuela['API Gravity'], api, lhv)
country_emissions = country_emissions.append(venezuela)

# Calculate subtotal CI to sort by for plot
country_emissions['Subtotal CI'] = country_emissions['Production Combustion CI'] + \
                                    country_emissions['Processing Combustion CI'] + \
                                    country_emissions['Total VFF CI'] + \
                                    country_emissions['Miscellaneous CI'] + \
                                    country_emissions['Offsite CI']
country_emissions = country_emissions.sort_values(by='Subtotal CI')

#########################################################
# generate emissions breakdown plot for each country (g/MJ)
#########################################################
N = len(country_emissions.index)
ind = np.arange(N)
expCI = country_emissions['Exploration Combustion CI']
dndCI = country_emissions['Drilling Combustion CI']
prodCI = country_emissions['Production Combustion CI']
procCI = country_emissions['Processing Combustion CI']
mainCI = country_emissions['Maintenance Combustion CI']
vffCI = country_emissions['Total VFF CI']
fCI = country_emissions['Flaring CI']
vfCI = country_emissions['Venting & Fugitives CI']
miscCI = country_emissions['Miscellaneous CI']
transCI = country_emissions['Transportation CI']
offsiteCI = country_emissions['Offsite CI']
totalCI = country_emissions['Total CI']
subtotalCI = country_emissions['Subtotal CI']

countryLabels = country_emissions.index.values.tolist()
#countryLabels[countryLabels.index('United States')] = 'USA\nUnconv.'

fig, ax = plt.subplots()
width = 0.75

prod_bar = ax.bar(ind, prodCI, width, label='Production', \
                  color=(7/255, 71/255, 77/255, 0.8))
proc_bar = ax.bar(ind, procCI, width, bottom=prodCI, \
                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
vff_bar = ax.bar(ind, vffCI, width, bottom=prodCI + procCI, \
            label='VFF', color=(1, 0, 0, 0.8))
misc_bar = ax.bar(ind, miscCI, width, bottom=prodCI + procCI + \
            vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
offsite_bar = ax.bar(ind, offsiteCI, width, bottom=prodCI + procCI + \
            vffCI + miscCI, label='Offsite', \
            color=(0,128/255,1,0.8))

ax.set_title('D18 Carbon Intensity Benchmarking')
ax.set_xticks(ind.tolist())
ax.set_xticklabels(countryLabels)
ax.set_ylabel('Carbon Intensity (gCO2e/MJ crude oil)')

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

handles, labels = ax.get_legend_handles_labels()
#handles = handles[4:] + handles[0:4]
#labels = labels[4:] + labels[0:4]
ax.legend(handles[::-1], labels[::-1], loc='center left', \
          bbox_to_anchor=(1, 0.5))

fig.savefig('D18_emissions_benchmarking.png', bbox_inches='tight', dpi=600)

#########################################################
# generate emissions breakdown plot for each country (te/bbl)
#########################################################
#N = len(country_emissions_bbl.index)
#ind = np.arange(N)
#expCI = country_emissions_bbl['Exploration Combustion CI']
#dndCI = country_emissions_bbl['Drilling Combustion CI']
#prodCI = country_emissions_bbl['Production Combustion CI']
#procCI = country_emissions_bbl['Processing Combustion CI']
#mainCI = country_emissions_bbl['Maintenance Combustion CI']
#vffCI = country_emissions_bbl['Total VFF CI']
#miscCI = country_emissions_bbl['Miscellaneous CI']
#transCI = country_emissions_bbl['Transportation CI']
#offsiteCI = country_emissions_bbl['Offsite CI']
#totalCI = country_emissions_bbl['Total CI']
#
#countryLabels = country_emissions.index.values.tolist()
#countryLabels[countryLabels.index('United States')] = 'USA\nUnconv.'
#countryLabels[countryLabels.index('Colombia')] = 'COL'
#countryLabels[countryLabels.index('Brazil')] = 'BRA'
#countryLabels[countryLabels.index('Argentina')] = 'ARG'
#countryLabels[countryLabels.index('Mexico')] = 'MEX'
#countryLabels[countryLabels.index('Ecuador')] = 'ECU'
#countryLabels[countryLabels.index('Venezuela')] = 'VEN'
#
#fig, ax = plt.subplots()
#width = 0.75
#exp_bar = ax.bar(ind, expCI, width, label='Exploration', \
#                 color=(45/255, 96/255, 179/255, 0.8))
#dnd_bar = ax.bar(ind, dndCI, width, label='Drilling', \
#                 color=(232/255, 159/255, 12/255, 0.8))
#prod_bar = ax.bar(ind, prodCI, width, bottom=dndCI, label='Production', \
#                  color=(7/255, 71/255, 77/255, 0.8))
#proc_bar = ax.bar(ind, procCI, width, bottom=dndCI + prodCI, \
#                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
#main_bar = ax.bar(ind, mainCI, width, bottom=dndCI + prodCI + procCI, \
#            label='Maintenance', color=(201/255, 195/255, 0, 0.8))
#vff_bar = ax.bar(ind, vffCI, width, bottom=dndCI + prodCI + procCI + \
#            mainCI, label='VFF', color=(255/255, 0, 0, 0.8))
#misc_bar = ax.bar(ind, miscCI, width, bottom=dndCI + prodCI + procCI + \
#            mainCI + vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
#trans_bar = ax.bar(ind, transCI, width, bottom=dndCI + prodCI + procCI + \
#            mainCI + vffCI + miscCI, label='Transport', \
#            color=(69/255,74/255,65/255,0.8))
#offsite_bar = ax.bar(ind, offsiteCI, width, bottom=dndCI + prodCI + procCI + \
#            mainCI + vffCI + miscCI + transCI, label='Offsite', \
#            color=(0,128/255,1,0.8))
#totalCI_line = ax.plot(ind, totalCI, 'kX-', label='Total CI')
#
#
#ax.set_title('Carbon Intensity Comparison')
#ax.set_xticks(ind.tolist())
#ax.set_xticklabels(countryLabels)
#ax.set_ylabel('Carbon Intensity (tCO2e/bbl crude oil)')
#
## Shrink current axis by 20%
#box = ax.get_position()
#ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])
#
#handles, labels = ax.get_legend_handles_labels()
#handles = handles[1:] + [handles[0]]
#labels = labels[1:] + [labels[0]]
#ax.legend(handles[::-1], labels[::-1], loc='center left', \
#          bbox_to_anchor=(1, 0.5))
#
#fig.savefig('emissions_breakdown_bbl.png', dpi=300)
