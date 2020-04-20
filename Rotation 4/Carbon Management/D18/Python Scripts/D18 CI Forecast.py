# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 12:31:23 2019

@author: 212566876
"""

#import seaborn as sns
import datetime
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import openpyxl
import pandas as pd
import numpy as np

yearloc = mdates.YearLocator()
monthloc = mdates.MonthLocator()
years_fmt = mdates.DateFormatter('%Y')

SMALL_SIZE = 8
MEDIUM_SIZE = 10
BIGGER_SIZE = 12

plt.rc('font', size=SMALL_SIZE)          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=SMALL_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=SMALL_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title

# load workbook for OPGEE GI scenario
wb = openpyxl.load_workbook('./D18 OPGEE Forecast GI.xlsm', data_only=True)
ws = wb['Results']

# read in needed columns for plto data
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
GOR = []            # gas oil ratio (scf/bbl)
WOR = []            # water oil ratio (bbl water/bbl oil)
WIR = []            # water injection ratio (bbl water/bbl oil)
api_gravity = []            # api gravity of crude
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []
flaring_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    api_gravity.append(ws.cell(row=34, column=col).value)
    GOR.append(ws.cell(row=46, column=col).value)
    WOR.append(ws.cell(row=47, column=col).value)
    WIR.append(ws.cell(row=48, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)
    flaring_CI.append(ws.cell(row=180, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1      
        

# convert lists to pandas dataframe
D18_data_GI = {'Country':countries, 'Field':fields, 'Production':production, \
        'GOR':GOR, 'WOR':WOR, 'WIR':WIR, \
        'API Gravity':api_gravity, 'Flaring (scf/bbl)':flaring_ratio, \
        'Flaring CI':flaring_CI, \
        'Exploration Combustion CI':exp_comb, \
        'Exploration VFF CI':exp_vff, \
        'Drilling Combustion CI':dnd_comb, \
        'Drilling VFF CI':dnd_vff, \
        'Production Combustion CI':prod_comb, \
        'Production VFF CI':prod_vff, \
        'Processing Combustion CI':proc_comb, \
        'Processing VFF CI':proc_vff, \
        'Maintenance Combustion CI':main_comb, \
        'Maintenance VFF CI':main_vff, \
        'Transportation CI':trans, \
        'Miscellaneous CI':misc, \
        'Offsite CI':offsite, \
        'Total CI':net_CI
    }


D18_emissions_GI = pd.DataFrame(D18_data_GI)
D18_emissions_GI['WIR'] = pd.to_numeric(D18_emissions_GI['WIR'], errors='coerce')
D18_emissions_GI['Total VFF CI'] = D18_emissions_GI['Exploration VFF CI'] + \
                            D18_emissions_GI['Drilling VFF CI'] + \
                            D18_emissions_GI['Production VFF CI'] + \
                            D18_emissions_GI['Processing VFF CI'] + \
                            D18_emissions_GI['Maintenance VFF CI']

# load workbook for OPGEE GM scenario
wb = openpyxl.load_workbook('./D18 OPGEE Forecast GM.xlsm', data_only=True)
ws = wb['Results']

# read in needed columns for plto data
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
GOR = []            # gas oil ratio (scf/bbl)
WOR = []            # water oil ratio (bbl water/bbl oil)
WIR = []            # water injection ratio (bbl water/bbl oil)
api_gravity = []            # api gravity of crude
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []
flaring_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    api_gravity.append(ws.cell(row=34, column=col).value)
    GOR.append(ws.cell(row=46, column=col).value)
    WOR.append(ws.cell(row=47, column=col).value)
    WIR.append(ws.cell(row=48, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)
    flaring_CI.append(ws.cell(row=180, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1      
        

# convert lists to pandas dataframe
D18_data_GM = {'Country':countries, 'Field':fields, 'Production':production, \
        'GOR':GOR, 'WOR':WOR, 'WIR':WIR, \
        'API Gravity':api_gravity, 'Flaring (scf/bbl)':flaring_ratio, \
        'Flaring CI':flaring_CI, \
        'Exploration Combustion CI':exp_comb, \
        'Exploration VFF CI':exp_vff, \
        'Drilling Combustion CI':dnd_comb, \
        'Drilling VFF CI':dnd_vff, \
        'Production Combustion CI':prod_comb, \
        'Production VFF CI':prod_vff, \
        'Processing Combustion CI':proc_comb, \
        'Processing VFF CI':proc_vff, \
        'Maintenance Combustion CI':main_comb, \
        'Maintenance VFF CI':main_vff, \
        'Transportation CI':trans, \
        'Miscellaneous CI':misc, \
        'Offsite CI':offsite, \
        'Total CI':net_CI
    }


D18_emissions_GM = pd.DataFrame(D18_data_GM)
D18_emissions_GM['WIR'] = pd.to_numeric(D18_emissions_GM['WIR'], errors='coerce')
D18_emissions_GM['Total VFF CI'] = D18_emissions_GM['Exploration VFF CI'] + \
                            D18_emissions_GM['Drilling VFF CI'] + \
                            D18_emissions_GM['Production VFF CI'] + \
                            D18_emissions_GM['Processing VFF CI'] + \
                            D18_emissions_GM['Maintenance VFF CI']

# load workbook for OPGEE FA scenario
wb = openpyxl.load_workbook('./D18 OPGEE Forecast FA.xlsm', data_only=True)
ws = wb['Results']

# read in needed columns for plto data
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
GOR = []            # gas oil ratio (scf/bbl)
WOR = []            # water oil ratio (bbl water/bbl oil)
WIR = []            # water injection ratio (bbl water/bbl oil)
api_gravity = []            # api gravity of crude
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []
flaring_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    api_gravity.append(ws.cell(row=34, column=col).value)
    GOR.append(ws.cell(row=46, column=col).value)
    WOR.append(ws.cell(row=47, column=col).value)
    WIR.append(ws.cell(row=48, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)
    flaring_CI.append(ws.cell(row=180, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1      
        

# convert lists to pandas dataframe
D18_data_FA = {'Country':countries, 'Field':fields, 'Production':production, \
        'GOR':GOR, 'WOR':WOR, 'WIR':WIR, \
        'API Gravity':api_gravity, 'Flaring (scf/bbl)':flaring_ratio, \
        'Flaring CI':flaring_CI, \
        'Exploration Combustion CI':exp_comb, \
        'Exploration VFF CI':exp_vff, \
        'Drilling Combustion CI':dnd_comb, \
        'Drilling VFF CI':dnd_vff, \
        'Production Combustion CI':prod_comb, \
        'Production VFF CI':prod_vff, \
        'Processing Combustion CI':proc_comb, \
        'Processing VFF CI':proc_vff, \
        'Maintenance Combustion CI':main_comb, \
        'Maintenance VFF CI':main_vff, \
        'Transportation CI':trans, \
        'Miscellaneous CI':misc, \
        'Offsite CI':offsite, \
        'Total CI':net_CI
    }


D18_emissions_FA = pd.DataFrame(D18_data_FA)
D18_emissions_FA['WIR'] = pd.to_numeric(D18_emissions_FA['WIR'], errors='coerce')
D18_emissions_FA['Total VFF CI'] = D18_emissions_FA['Exploration VFF CI'] + \
                            D18_emissions_FA['Drilling VFF CI'] + \
                            D18_emissions_FA['Production VFF CI'] + \
                            D18_emissions_FA['Processing VFF CI'] + \
                            D18_emissions_FA['Maintenance VFF CI']

#########################################################
# generate emissions breakdown plot for each year (g/MJ)
#########################################################
N = len(D18_emissions_GI.index)
ind = np.arange(N)
totalCI_GI = D18_emissions_GI['Total CI']
totalCI_GM = D18_emissions_GM['Total CI']
totalCI_FA = D18_emissions_FA['Total CI']

years = D18_emissions_GI['Field'].tolist()
x_val = [datetime.date(year, 1, 1) for year in years]

fig, ax = plt.subplots()

# format the ticks
ax.xaxis.set_major_locator(yearloc)
ax.xaxis.set_major_formatter(years_fmt)
#ax.xaxis.set_minor_locator(monthloc)

ax.plot(x_val, totalCI_GI, 'b-v', label='Gas Injection')
ax.plot(x_val, totalCI_GM, 'g-d', label='Gas Monetization')
ax.plot(x_val, totalCI_FA, 'r-o', label='Continue Flaring')


ax.set_title('D18 Carbon Intensity Forecast 2019-36')
#ax.set_xticks(ind.tolist())
#ax.set_xticklabels(yearLabels)
ax.set_ylabel('Carbon Intensity\n(gCO2e/MJ crude oil)')

# Shrink current axis by 20%
#box = ax.get_position()
#ax.set_position([box.x0, box.y0, box.width * 0.7, box.height])
#ax.legend(loc='center left', bbox_to_anchor=(1, 0.5))
ax.legend()
fig.autofmt_xdate()
fig.savefig('D18_emissions_forecast_2019_to_2036.png', bbox_inches='tight', dpi=600)
