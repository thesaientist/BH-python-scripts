#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Apr 17 10:51:14 2020

@author: saiuppati

Library of functions to conduct input and output database updates
"""

import os
import matplotlib as mpl
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np
from fuzzywuzzy import process     # module for string matching subroutine

# matplotlib figure resolution
plt.rcParams["figure.dpi"] = 300
mpl.rcParams.update({'font.size': 12})
mpl.rc('xtick', labelsize=12)
mpl.rc('ytick', labelsize=12)

class ciDatabase:

    # initial method / constructor
    def __init__(self, name=None):
        if name is None:
            name = 'Science Paper CI Dataset'
        self.name = name
        self.set_source_database()
        self.set_output_excel()

    # get name of database (ACCESSOR)
    def get_name(self):
        print(self.name)

    # set output Excel worksheet
    def set_output_excel(self, excelFilePath=None):
        if excelFilePath is None:
            cwd = os.getcwd()
            excelFilePath = os.path.join(cwd, \
                            'Carbon Intensity - Global Dataset of Fields.xlsx')
        self.outputExcelPath = excelFilePath

        # set output excel workbook and sheet objects from openpyxl
        try:
            self.wbo = openpyxl.load_workbook(excelFilePath)
            self.wso = self.wbo['CI Database']
        except FileNotFoundError:
            print('Output excel file not found!')

    # get output Excel worksheet name (ACCESSOR)
    def get_output_excel_path(self):
        print(self.outputExcelPath)
        return self.outputExcelPath

    # set source database (by default, it's Rystad database spreadsheet in current working directory)
    def set_source_database(self, databaseFilePath=None):
        if databaseFilePath is None:
            cwd = os.getcwd()
            databaseFilePath = os.path.join(cwd, \
                                'Carbon Intensity - Rystad Export 33 Countries.xlsx')
        self.sourceDBPath = databaseFilePath

        # set source excel workbook and sheet objects from openpyxl
        try:
            self.wbi = openpyxl.load_workbook(databaseFilePath, data_only=True)
            self.wsi = self.wbi['Categorization']
        except FileNotFoundError:
            print('Source database excel file not found!')

    # get source database (ACCESSOR)
    def get_source_database_path(self):
        print(self.sourceDBPath)
        return self.sourceDBPath

    # create dataframe that contains data from source database that's properly
    # organized. try loading sourceData.pkl file, and if not available, call
    # extract_data_from_source() function to create sourceData.pkl file and dataframe
    def create_sourceData_dataframe(self, boolUpdate=False):
        # if source database has been recently updated...then extract and create .pkl file
        if boolUpdate:
            print('\nIt looks like the source database has been recently updated' +\
                  ' , so please wait while source dataframe is generated...\n')
            self.extract_data_from_source()
        else:
            # try loading sourceData.pkl and catch error
            try:
                self.sourceData = pd.read_pickle('sourceData.pkl')
                print('\nPickle file sourceData.pkl found, which was used' +\
                      ' to load source dataframe.\n')
            except FileNotFoundError:
                print('\nPickle file sourceData.pkl was not found in the current '+\
                      'working directory. Source data will be extracted from '+\
                      'source database at the default provided path. ' +\
                      'Please wait...\n')
                self.extract_data_from_source()
        # Task finished message
        print('\nSource dataframe has been created!\n')

    # extract data from source database into a properly organized pandas df
    def extract_data_from_source(self):

        # set input database excel sheet that was initialized in constructor
        wsi = self.wsi

        # read in columns from Rystad dataset (input workbook)
        country = []      # names of countries
        field = []         # names of fields
        basin = []
        block = []      # Main original contract / block
        province = []
        prod_type = []  # oil/gas/gas-condensate ('Field Type Category')
        unconv_detail = []  # 'Uncoventional detail' i.e. extra heavy oil, oil sands etc.
        archetype = []  # 'Supply Segment Group' i.e. Offshore shelf, other onshore etc.
        start_year = [] # start up year
        field_size = [] # amount of resources
        status = []     # producing/abandoned etc
        operator = []   # operating company name
        operator_type = []  # i.e. NOC v INOC etc
        company_interest = []   # ownership splits (names and percentages)
        production = [] # current production rate (kbbl/d)
        depth = []      # water depth (ft) if negative / land elevation (ft) if positive

        # collect data until last existing column
        row = 2
        while True:
            country.append(wsi.cell(row=row, column=2).value)
            field.append(wsi.cell(row=row, column=3).value)
            basin.append(wsi.cell(row=row, column=4).value)
            block.append(wsi.cell(row=row, column=5).value)
            province.append(wsi.cell(row=row, column=6).value)
            prod_type.append(wsi.cell(row=row, column=7).value)
            unconv_detail.append(wsi.cell(row=row, column=8).value)
            archetype.append(wsi.cell(row=row, column=9).value)
            start_year.append(wsi.cell(row=row, column=10).value)
            field_size.append(wsi.cell(row=row, column=11).value)
            status.append(wsi.cell(row=row, column=12).value)
            operator.append(wsi.cell(row=row, column=13).value)
            operator_type.append(wsi.cell(row=row, column=14).value)
            company_interest.append(wsi.cell(row=row, column=15).value)
            depth.append(wsi.cell(row=row, column=16).value)
            production.append(wsi.cell(row=row, column=17).value)

            nextField = wsi.cell(row=row+1, column=3).value
            if not nextField:
                break
            else:
                row += 1

        # convert some data types (string to number/integer or units)
        for i in range(len(field)):
            start_year[i] = int(start_year[i])
            production[i] = production[i] * 1000    # kbbl/d to bbl/d
            if depth[i] == '':
                depth[i] = 0.0
            else:
                depth[i] = float(depth[i])

        # separate ownership shares into separate columns, for each field
        op_int = np.zeros(len(company_interest))
        companies = [['', '', '', '', '', ''] for x in range(len(company_interest))]
        comp_int = np.zeros((len(company_interest), 6))

        for i in range(len(company_interest)):
            strng = company_interest[i]
            n = 0   # which column in array to store
            while True:
                if n>5:
                    break
                pos = strng.find('%)')
                substrng = strng[:pos+2]
                pos = substrng.rfind(' (')
                company_name = strng[:pos]
                pos2 = strng.find('%)')
                company_share = float(strng[pos+2:pos2])
                if '*' in company_name:
                    op_int[i] = company_share
                else:
                    companies[i][n] = company_name
                    comp_int[i, n] = company_share
                    n += 1

                # check if first / remaining company is the last owner
                pos = strng.find(';')
                if pos == -1:
                    break
                else:
                    strng = strng.replace(strng[:pos+2], '')

        # # convert companies list of lists to string array
        companies = np.array(companies)

        # convert pandas dataframe
        data = {'Country':country, 'Field':field, 'Basin':basin, \
                'Block':block, \
                'Province':province, \
                'Field Type':prod_type, \
                'Unconventional Detail':unconv_detail, \
                'Archetype':archetype, \
                'Start Year':start_year, \
                'Field Size':field_size, \
                'Field Status':status, \
                'Operator':operator, \
                'Operator Type':operator_type, \
                'Operator Interest':op_int, \
                'Company 2': companies[:,0], \
                'Company 2 Interest': comp_int[:,0], \
                'Company 3': companies[:,1], \
                'Company 3 Interest': comp_int[:,1], \
                'Company 4': companies[:,2], \
                'Company 4 Interest': comp_int[:,2], \
                'Company 5': companies[:,3], \
                'Company 5 Interest': comp_int[:,3], \
                'Company 6': companies[:,4], \
                'Company 6 Interest': comp_int[:,4], \
                'Company 7': companies[:,5], \
                'Company 7 Interest': comp_int[:,5], \
                'Depth': depth, \
                'Production':production
            }
        sourceData = pd.DataFrame(data)

        # replace company interest percentages to NaN when it's zeros
        cols = ["Operator Interest","Company 2 Interest", "Company 3 Interest", \
                "Company 4 Interest", "Company 5 Interest", "Company 6 Interest", \
                "Company 7 Interest"]
        sourceData[cols] = sourceData[cols].replace({0:np.nan})

        # set source data dataframe attribute to class
        self.sourceData = sourceData

        # classify archetypes for source data
        self.classify_archetypes()

        # save archetype categorized dataframe to load for repeated use
        self.sourceData.to_pickle('sourceData.pkl')

    # access categorized source data dataframe (ACCESSOR)
    def get_source_dataframe(self):
        return self.sourceData

    # manage archetype classification of source data
    def classify_archetypes(self):

        # source dataframe
        df = self.sourceData

        # offshore archetype classification based on depth
        shelfFilter = (df['Depth'] < 0) & (df['Depth'] > -5000)
        deepwaterFilter = (df['Depth'] <= -5000) & (df['Depth'] > -10000)
        ultradeepFilter = df['Depth'] <= -10000
        df.loc[shelfFilter, 'Archetype'] = 'Offshore shelf'
        df.loc[deepwaterFilter, 'Archetype'] = 'Offshore deepwater'
        df.loc[ultradeepFilter, 'Archetype'] = 'Offshore ultra deepwater'
        offshoreFilter = df['Archetype'].str.contains('Offshore')

        # if 'arctic' then reclassify
        arcticFilter = df['Unconventional Detail'] == 'Arctic'
        df.loc[arcticFilter, 'Archetype'] =  'Arctic ' + \
            df.loc[arcticFilter, 'Archetype'].astype(str)

        # if 'extra heavy oil' then reclassify
        extraHeavyFilter = df['Unconventional Detail'] == 'Extra heavy oil'
        onshoreFilter = ~offshoreFilter
        df.loc[extraHeavyFilter & offshoreFilter, 'Archetype'] = \
            df.loc[extraHeavyFilter & offshoreFilter, 'Archetype'].astype(str) + \
                ' extra heavy oil'
        df.loc[extraHeavyFilter & onshoreFilter, 'Archetype'] = 'Extra heavy oil'

        # NAM 'shale/tight' classification
        shaleTightFilter = df['Archetype'] == 'Shale/Tight oil'
        NAM_countries = ['United States', 'Canada', 'Mexico']
        NAM_filter = df['Country'].isin(NAM_countries)
        df.loc[shaleTightFilter & NAM_filter, 'Archetype'] = 'NAM Shale/Tight oil'
        df.loc[shaleTightFilter & ~NAM_filter, 'Archetype'] = 'Other Shale/Tight oil'

        # if 'shale or tight' and offshore then reclassify
        shaleTightUNCFilter = (df['Unconventional Detail'].str.contains('Shale')) | \
                                (df['Unconventional Detail'].str.contains('Tight'))
        df.loc[shaleTightUNCFilter & offshoreFilter, 'Archetype'] = \
            df.loc[shaleTightUNCFilter & offshoreFilter, 'Archetype'].astype(str) + \
                ' Shale/Tight oil'

        # Other onshore classification into country/region based classification
        otherOnshoreFilter = df['Archetype'] == 'Other Onshore'
        ME_countries = ['Akrotiri and Dhekelia', \
                        'Bahrain', \
                        'Cyprus', \
                        'Egypt', \
                        'Iran', \
                        'Iraq', \
                        'Israel', \
                        'Jordan', \
                        'Kuwait', \
                        'Lebanon', \
                        'Oman', \
                        'Palestine', \
                        'Qatar', \
                        'Saudi Arabia', \
                        'Turkey', \
                        'United Arab Emirates', \
                        'UAE', \
                        'Yemen'
                        ]
        ME_filter = df['Country'].isin(ME_countries)
        russiaFilter = df['Country'] == 'Russia'
        df.loc[otherOnshoreFilter & ME_filter, 'Archetype'] = 'Onshore Middle East'
        df.loc[otherOnshoreFilter & russiaFilter, 'Archetype'] = 'Onshore Russia'
        df.loc[otherOnshoreFilter & ~ME_filter & ~russiaFilter, 'Archetype'] = \
            'Onshore RoW'

        # if 'coal bed methane' then append this type to end of Archetype
        cbmFilter = df['Unconventional Detail'] == 'Coalbed methane'
        df.loc[cbmFilter, 'Archetype'] = df.loc[cbmFilter, 'Archetype'].astype(str) + \
            ' coalbed methane'

        # if 'oil shale' then append this type to end of Archetype
        oilShaleFilter = df['Unconventional Detail'].str.contains('Oil shale')
        df.loc[oilShaleFilter, 'Archetype'] = df.loc[oilShaleFilter, 'Archetype'].astype(str) + \
            ' oil shale'

        # update sourceData dataframe attribute of class with archetype categorization
        self.sourceData = df

    # select specific columns of data from CI database excel to import into
    # pandas dataframe
    def create_ci_dataframe(self, boolUpdate=False):
        # if CI database needs to be updated...then call load_output_excel()
        # to populate the excel with latest categorization data and save workbook
        if boolUpdate:
            print('\nIt looks like the CI database needs to be updated' +\
                  ' , so please wait while the excel is re-populated and ' +\
                  ' pandas dataframe is generated...\n')
            self.load_output_excel(True)
        else:
            # try loading ci_database.pkl and catch error
            try:
                self.cidf = pd.read_pickle('ci_database.pkl')
                print('\nPickle file ci_database.pkl found, which was used' +\
                      ' to load CI dataframe.\n')
            except FileNotFoundError:
                print('\nPickle file ci_database.pkl was not found in the current '+\
                      'working directory. CI database excel will be loaded from '+\
                          'source database and new ci_database.pkl file created.\n')
                self.load_output_excel(True)
        print('CI dataframe has been created!')

    # access CI database dataframe (ACCESSOR)
    def get_ci_dataframe(self):
        return self.cidf

    # populate CI database output excel sheet using source dataframe
    # and matching fields and then create CI database dataframe
    def load_output_excel(self, clearPrev=True):
        # define source dataframe and output worksheet
        self.create_sourceData_dataframe()
        sourceData = self.sourceData
        wso = self.wso

        # clear existing data from previous populations if clearPrev is True
        if clearPrev:
            r = 2
            while True:
                for c in range(4, 28):
                    wso.cell(row=r, column=c).value = None
                if wso.cell(row=r+1, column=3).value is None:
                    break
                r += 1

        # Loop to populate output CI database excel sheet, with field matching
        row = 2
        while True:
            # set searchCountry and searchField for filtering source dataframe
            searchCountry = wso.cell(row=row, column=2).value
            searchField = wso.cell(row=row, column=3).value

            # modify some searchCountry names to match source dataframe
            if searchCountry == 'Republic of Congo':
                searchCountry = 'Congo'
            if searchCountry == 'Russian Federation':
                searchCountry = 'Russia'
            if searchCountry == 'United Arab Emirates':
                searchCountry = 'UAE'

            # find the subset of data for the right country
            isCountry = sourceData['Country'].str.contains(searchCountry)
            subset = sourceData[isCountry]

            # find the rows of data that match the field name (fuzzy string matching)
            countryFields = subset['Field'].tolist()
            bestMatch = process.extractOne(searchField, countryFields)
            # high score of 90 or more required to satisfy field name match
            if bestMatch[1] >= 90:
                matchedField = bestMatch[0]
                isField = subset['Field'] == matchedField
                subset2 = subset[isField]
            else:
                # because the score is lower than 90, subset2 will be empty
                subset2 = subset[subset['Field']==searchField]

            # check if subset2 is empty, otherwise populate columns in output sheet
            isEmpty = subset2.empty
            if isEmpty:
                row += 1
                continue
            else:
                wso.cell(row=row, column=4).value = subset2.iloc[0]['Basin']
                wso.cell(row=row, column=5).value = subset2.iloc[0]['Block']
                wso.cell(row=row, column=6).value = subset2.iloc[0]['Province']
                wso.cell(row=row, column=7).value = subset2.iloc[0]['Field Type']
                archetyp = subset2.iloc[0]['Archetype']
                if 'Offshore' in archetyp:
                    landvsea = 'Offshore'
                else:
                    landvsea = 'Onshore'
                wso.cell(row=row, column=8).value = landvsea
                wso.cell(row=row, column=9).value = subset2.iloc[0]['Depth']
                wso.cell(row=row, column=10).value = archetyp
                wso.cell(row=row, column=11).value = subset2.iloc[0]['Start Year']
                wso.cell(row=row, column=12).value = subset2.iloc[0]['Operator']
                wso.cell(row=row, column=13).value = subset2.iloc[0]['Operator Type']
                wso.cell(row=row, column=14).value = subset2.iloc[0]['Operator Interest']
                wso.cell(row=row, column=15).value = subset2.iloc[0]['Company 2']
                wso.cell(row=row, column=16).value = subset2.iloc[0]['Company 2 Interest']
                wso.cell(row=row, column=17).value = subset2.iloc[0]['Company 3']
                wso.cell(row=row, column=18).value = subset2.iloc[0]['Company 3 Interest']
                wso.cell(row=row, column=19).value = subset2.iloc[0]['Company 4']
                wso.cell(row=row, column=20).value = subset2.iloc[0]['Company 4 Interest']
                wso.cell(row=row, column=21).value = subset2.iloc[0]['Company 5']
                wso.cell(row=row, column=22).value = subset2.iloc[0]['Company 5 Interest']
                wso.cell(row=row, column=23).value = subset2.iloc[0]['Company 6']
                wso.cell(row=row, column=24).value = subset2.iloc[0]['Company 6 Interest']
                wso.cell(row=row, column=25).value = subset2.iloc[0]['Company 7']
                wso.cell(row=row, column=26).value = subset2.iloc[0]['Company 7 Interest']
                wso.cell(row=row, column=27).value = subset2.iloc[0]['Production']

            # check if next row is not empty; if empty, then break out of loop because reached end
            nextField = wso.cell(row=row+1, column=3).value
            if not nextField:
                break
            else:
                row += 1

        # save workbook
        self.wbo.save(self.outputExcelPath)

        # convert from excel into pandas dataframe and save local .pkl file
        self.ci_excel_to_df()

    # convert CI database data from populated excel sheet to pandas dataframe
    # and save local ci_database.pkl file for use later
    def ci_excel_to_df(self):
        ciData = pd.read_excel(self.outputExcelPath, sheet_name='CI Database')
        self.cidf = pd.DataFrame(ciData, columns=['Country','Field','Basin','Block / Contract', \
                                            'Province','Production Type','On/Offshore', \
                                            'Depth/Elevation (ft)','Archetype', \
                                            'Production Start Year', 'Company 1/Operator', \
                                            'Operator Type','Company 1/Operator Interest (%)', \
                                            'Company 2','Company 2 Interest (%)','Company 3', \
                                            'Company 3 Interest (%)','Company 4', \
                                            'Company 4 Interest (%)','Company 5', \
                                            'Company 5 Interest (%)','Company 6', \
                                            'Company 6 Interest (%)','Company 7', \
                                            'Company 7 Interest (%)','Production (bbl/d)', \
                                            'Exploration','Drilling', \
                                            'Production','Processing', \
                                            'VFF', \
                                            'Miscellaneous','Transport', \
                                            'Offsite', \
                                            'CI (gCO2e/MJ)', \
                                            'Water flooding', 'Original Input WOR', 'OPGEE Adjusted WOR', \
                                            'Original Input GOR (scf/bbl)', 'OPGEE Adjusted GOR (scf/bbl)', \
                                            'OPGEE Adjusted Flaring Ratio (scf/bbl)', \
                                            ])

        # save CI database dataframe as pickle file
        self.cidf.to_pickle('ci_database.pkl')

class ciSubset:

    # initial method/constructor
    def __init__(self, name, df):
        self.df = df
        self.name = name
        self.vwavg_ci()

    # determine volume weighted average of all entries by production and
    # return average CI and components in a dataframe
    def vwavg_ci(self, weight='Production (bbl/d)'):
        df = self.df
        qtys = ['Exploration', 'Drilling', 'Production', 'Processing', 'VFF', \
                'Transport', 'Offsite', 'CI (gCO2e/MJ)']
        data = dict.fromkeys(qtys)
        for qty in qtys:
            data[qty] = self.wavg(df, qty, weight)
        self.avg_df = pd.DataFrame(data=data, index=[self.name])

    # determine volume weighted averages of subsets of provided df, weighted
    # by production and return average CI and components in a dataframe
    def groupby_vwavg_ci(self, grp, weight='Production (bbl/d)'):
        df = self.df
        qtys = ['Exploration', 'Drilling', 'Production', 'Processing', 'VFF', \
                'Transport', 'Offsite', 'CI (gCO2e/MJ)']
        data = dict.fromkeys(qtys)
        for qty in qtys:
            data[qty] = df.groupby(grp).apply(self.wavg, qty, weight)
        avg_df1 = self.avg_df
        avg_df2 = pd.DataFrame(data=data)
        frames = [avg_df1, avg_df2]
        self.avg_df = pd.concat(frames)

    # access volume weighted averages calculated thus far for subset (ACCESSOR)
    def get_averages(self):
        return self.avg_df

    # weighted average function to call in other member functions
    def wavg(self, group, avg_name, weight_name):
        d = group[avg_name]
        w = group[weight_name]
        try:
            return (d * w).sum() / w.sum()
        except ZeroDivisionError:
            return d.mean()

    # method to generate standard stacked barplot for CI averages
    def plot_ci(self, option=0, subset=None):
        # option determines which components of CI get plotted
        #   0 = Exp., Drill., Prod., Proc., VFF., Transp., Offsite (total CI gets omitted)
        #   1 = Prod., Proc., VFF., Offsite.
        #   2 = Total CI ONLY
        # subset (if provided) means an alternate dataframe will be plotted
        avg_df = self.avg_df.sort_values(by=['CI (gCO2e/MJ)'])
        if subset is not None:
            avg_df = subset
        # choose portions of dataframe and custom colors based on option for plot
        if option == 0:
            avg_df = avg_df.iloc[:,:-1]
            customCIColors = [(45/255, 96/255, 179/255, 0.8), \
                            (232/255, 159/255, 12/255, 0.8), \
                            (7/255, 71/255, 77/255, 0.8), \
                            (247/255, 208/255, 12/255, 0.8), \
                            (225/255, 10/255, 0, 0.8), \
                            (69/255,74/255,65/255,0.8), \
                            (0,128/255,1,0.8), \
                            ]
        elif option == 1:
            avg_df = avg_df.loc[:,['Production', 'Processing', 'VFF', 'Offsite']]
            customCIColors = [\
                            (7/255, 71/255, 77/255, 0.8), \
                            (247/255, 208/255, 12/255, 0.8), \
                            (225/255, 10/255, 0, 0.8), \
                            (0,128/255,1,0.8), \
                            ]
        elif option == 2:
            avg_df = avg_df.iloc[:,-1]
            customCIColors = [\
                            (225/255, 10/255, 0, 0.8), \
                            ]
        ax = avg_df.plot(kind='barh', stacked=True, color=customCIColors)
        ax.set_xlabel('gCO2e/MJ')
        self.avg_ci_plot = ax
        return ax
