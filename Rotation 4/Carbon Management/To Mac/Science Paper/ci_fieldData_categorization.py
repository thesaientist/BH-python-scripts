# -*- coding: utf-8 -*-
"""
Created on Mon Apr 6 2020

@author: 212566876
"""

#import seaborn as sns
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime

# start timer
startTime = datetime.now()

# load input and output workbooks
# wbi = openpyxl.load_workbook('./Carbon Intensity - Rystad Export 33 Countries.xlsx', data_only=True)
# wsi = wbi['Categorization']
wbo = openpyxl.load_workbook('./Carbon Intensity - Global Dataset of Fields - Copy 3.xlsx')
wso = wbo['Sheet1']

# # read in columns from Rystad dataset (input workbook)
# country = []      # names of countries
# field = []         # names of fields
# basin = []
# block = []      # Main original contract / block
# province = []
# prod_type = []  # oil/gas/gas-condensate ('Field Type Category')
# archetype = []  # 'Supply Segment Group' i.e. Offshore shelf, other onshore etc.
# start_year = [] # start up year
# field_size = [] # amount of resources
# status = []     # producing/abandoned etc
# operator = []   # operating company name
# operator_type = []  # i.e. NOC v INOC etc
# company_interest = []   # ownership splits (names and percentages)
# production = [] # current production rate (kbbl/d)
# depth = []      # water depth (ft) if negative / land elevation (ft) if positive
#
# # collect data until last existing column
# row = 2
# while True:
#     country.append(wsi.cell(row=row, column=2).value)
#     field.append(wsi.cell(row=row, column=3).value)
#     basin.append(wsi.cell(row=row, column=4).value)
#     block.append(wsi.cell(row=row, column=5).value)
#     province.append(wsi.cell(row=row, column=6).value)
#     prod_type.append(wsi.cell(row=row, column=7).value)
#     archetype.append(wsi.cell(row=row, column=8).value)
#     start_year.append(wsi.cell(row=row, column=9).value)
#     field_size.append(wsi.cell(row=row, column=10).value)
#     status.append(wsi.cell(row=row, column=11).value)
#     operator.append(wsi.cell(row=row, column=12).value)
#     operator_type.append(wsi.cell(row=row, column=13).value)
#     company_interest.append(wsi.cell(row=row, column=14).value)
#     depth.append(wsi.cell(row=row, column=15).value)
#     production.append(wsi.cell(row=row, column=16).value)
#
#     nextField = wsi.cell(row=row+1, column=3).value
#     if not nextField:
#         break
#     else:
#         row += 1
#
# # convert some data types (string to number/integer or units)
# for i in range(len(field)):
#     start_year[i] = int(start_year[i])
#     production[i] = production[i] * 1000    # kbbl/d to bbl/d
#     if depth[i] == '':
#         depth[i] = 0.0
#     else:
#         depth[i] = float(depth[i])
#
# # correct offshore archetype classification
# for i in range(len(archetype)):
#     if depth[i] > -5000 and depth[i] < 0:
#         archetype[i] = 'Offshore shelf'
#     elif depth[i] > -10000 and depth[i] <= -5000:
#         archetype[i] = 'Offshore deepwater'
#     elif depth[i] <= -10000:
#         archetype[i] = 'Offshore ultra deepwater'
#
# # separate ownership shares into separate columns, for each field
# op_int = np.zeros(len(company_interest))
# companies = [['', '', '', '', '', ''] for x in range(len(company_interest))]
# comp_int = np.zeros((len(company_interest), 6))
#
# for i in range(len(company_interest)):
#     strng = company_interest[i]
#     n = 0   # which column in array to store
#     while True:
#         if n>5:
#             break
#         pos = strng.find('%)')
#         substrng = strng[:pos+2]
#         pos = substrng.rfind(' (')
#         company_name = strng[:pos]
#         pos2 = strng.find('%)')
#         company_share = float(strng[pos+2:pos2])
#         if '*' in company_name:
#             op_int[i] = company_share
#         else:
#             companies[i][n] = company_name
#             comp_int[i, n] = company_share
#             n += 1
#
#         # check if first / remaining company is the last owner
#         pos = strng.find(';')
#         if pos == -1:
#             break
#         else:
#             strng = strng.replace(strng[:pos+2], '')
#
# # # convert companies list of lists to string array
# companies = np.array(companies)
#
# # convert pandas dataframe
# data = {'Country':country, 'Field':field, 'Basin':basin, \
#         'Block':block, \
#         'Province':province, \
#         'Field Type':prod_type, \
#         'Archetype':archetype, \
#         'Start Year':start_year, \
#         'Field Size':field_size, \
#         'Field Status':status, \
#         'Operator':operator, \
#         'Operator Type':operator_type, \
#         'Operator Interest':op_int, \
#         'Company 2': companies[:,0], \
#         'Company 2 Interest': comp_int[:,0], \
#         'Company 3': companies[:,1], \
#         'Company 3 Interest': comp_int[:,1], \
#         'Company 4': companies[:,2], \
#         'Company 4 Interest': comp_int[:,2], \
#         'Company 5': companies[:,3], \
#         'Company 5 Interest': comp_int[:,3], \
#         'Company 6': companies[:,4], \
#         'Company 6 Interest': comp_int[:,4], \
#         'Company 7': companies[:,5], \
#         'Company 7 Interest': comp_int[:,5], \
#         'Depth': depth, \
#         'Production':production
#     }
# rystadData = pd.DataFrame(data)
#
# # replace company interest percentages to NaN when it's zeros
# cols = ["Operator Interest","Company 2 Interest", "Company 3 Interest", \
#         "Company 4 Interest", "Company 5 Interest", "Company 6 Interest", \
#         "Company 7 Interest"]
# rystadData[cols] = rystadData[cols].replace({0:np.nan})
#
# # save rystadData dataframe to load for repeated use
# rystadData.to_pickle('ciRystadData33CountriesCorrected.pkl')

# load rystadData dataframe from saved ciRystadData33Countries.pkl file
rystadData = pd.read_pickle('ciRystadData33CountriesCorrected.pkl')

# populate data into output sheet until last field
row = 2
while True:
    searchCountry = wso.cell(row=row, column=2).value
    searchField = wso.cell(row=row, column=3).value

    if searchCountry == 'Republic of Congo':
        searchCountry = 'Congo'
    if searchCountry == 'Russian Federation':
        searchCountry = 'Russia'
    if searchCountry == 'United Arab Emirates':
        searchCountry = 'UAE'

    # find the subset of data for the right country
    isCountry = rystadData['Country'].str.contains(searchCountry)
    subset = rystadData[isCountry]

    # find the rows of data that match the field name
    countryFields = subset['Field'].tolist()
    from fuzzywuzzy import process
    bestMatch = process.extractOne(searchField, countryFields)
    if bestMatch[1] >= 90:
        matchedField = bestMatch[0]
        isField = subset['Field'] == matchedField
        subset2 = subset[isField]
    else:
        subset2 = subset[subset['Field']==searchField]

    # check if subset2 is empty, otherwise populate columns in output sheet
    isEmpty = subset2.empty
    if isEmpty:
        row += 1
        continue
    else:
        wso.cell(row=row, column=4).value = subset2.iloc[0]['Basin']
        wso.cell(row=row, column=5).value = subset2.iloc[0]['Block']
        wso.cell(row=row, column=6).value = subset2.iloc[0]['Province']
        wso.cell(row=row, column=7).value = subset2.iloc[0]['Field Type']
        archetyp = subset2.iloc[0]['Archetype']
        if 'Offshore' in archetyp:
            landvsea = 'Offshore'
        else:
            landvsea = 'Onshore'
        wso.cell(row=row, column=8).value = landvsea
        wso.cell(row=row, column=9).value = subset2.iloc[0]['Depth']
        wso.cell(row=row, column=10).value = archetyp
        wso.cell(row=row, column=11).value = subset2.iloc[0]['Start Year']
        wso.cell(row=row, column=12).value = subset2.iloc[0]['Operator']
        wso.cell(row=row, column=13).value = subset2.iloc[0]['Operator Type']
        wso.cell(row=row, column=14).value = subset2.iloc[0]['Operator Interest']
        wso.cell(row=row, column=15).value = subset2.iloc[0]['Company 2']
        wso.cell(row=row, column=16).value = subset2.iloc[0]['Company 2 Interest']
        wso.cell(row=row, column=17).value = subset2.iloc[0]['Company 3']
        wso.cell(row=row, column=18).value = subset2.iloc[0]['Company 3 Interest']
        wso.cell(row=row, column=19).value = subset2.iloc[0]['Company 4']
        wso.cell(row=row, column=20).value = subset2.iloc[0]['Company 4 Interest']
        wso.cell(row=row, column=21).value = subset2.iloc[0]['Company 5']
        wso.cell(row=row, column=22).value = subset2.iloc[0]['Company 5 Interest']
        wso.cell(row=row, column=23).value = subset2.iloc[0]['Company 6']
        wso.cell(row=row, column=24).value = subset2.iloc[0]['Company 6 Interest']
        wso.cell(row=row, column=25).value = subset2.iloc[0]['Company 7']
        wso.cell(row=row, column=26).value = subset2.iloc[0]['Company 7 Interest']
        wso.cell(row=row, column=27).value = subset2.iloc[0]['Production']

    nextField = wso.cell(row=row+1, column=3).value
    if not nextField:
        break
    else:
        row += 1

# save workbook
wbo.save('./Carbon Intensity - Global Dataset of Fields - Copy 3.xlsx')

# print time taken to execute script
print(datetime.now() - startTime)

#
# # generate emissions breakdown plot for each field
# N = len(emissions.index)
# ind = np.arange(N)
# dndCI = emissions['Drilling Combustion CI']
# prodCI = emissions['Production Combustion CI']
# procCI = emissions['Processing Combustion CI']
# mainCI = emissions['Maintenance Combustion CI']
# vffCI = emissions['Total VFF CI']
# miscCI = emissions['Miscellaneous CI']
# transCI = emissions['Transportation CI']
# offsiteCI = emissions['Offsite CI']
#
# fieldLabels = emissions['Field'].tolist()
# # legendLabels = ['Drilling', 'Production', 'Processing', 'Maintenance', 'VFF', \
# #                 'Misc.', 'Transport']
#
# fig, ax = plt.subplots()
# width = 0.75
# # color=(45/255, 96/255, 179/255, 0.8)  # color for exploration bar if ever added
# dnd_bar = ax.bar(ind, dndCI, width, label='Drilling', \
#                  color=(232/255, 159/255, 12/255, 0.8))
# prod_bar = ax.bar(ind, prodCI, width, bottom=dndCI, label='Production', \
#                   color=(7/255, 71/255, 77/255, 0.8))
# proc_bar = ax.bar(ind, procCI, width, bottom=dndCI + prodCI, \
#                   label='Processing', color=(247/255, 208/255, 12/255, 0.8))
# main_bar = ax.bar(ind, mainCI, width, bottom=dndCI + prodCI + procCI, \
#             label='Maintenance', color=(201/255, 195/255, 0, 0.8))
# vff_bar = ax.bar(ind, vffCI, width, bottom=dndCI + prodCI + procCI + \
#             mainCI, label='VFF', color=(225/255, 10/255, 0, 0.8))
# misc_bar = ax.bar(ind, miscCI, width, bottom=dndCI + prodCI + procCI + \
#             mainCI + vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
# trans_bar = ax.bar(ind, transCI, width, bottom=dndCI + prodCI + procCI + \
#             mainCI + vffCI + miscCI, label='Transport', \
#             color=(69/255,74/255,65/255,0.8))
# offsite_bar = ax.bar(ind, offsiteCI, width, bottom=dndCI + prodCI + procCI + \
#             mainCI + vffCI + miscCI + transCI, label='Offsite', \
#             color=(0,128/255,1,0.8))
#
# ax.set_title('GHG Emissions Intensity Comparison')
# ax.set_xticks(ind.tolist())
# ax.set_xticklabels(fieldLabels)
# ax.set_ylabel('GHG Intensity (gCO2e/MJ crude oil)')
# # ax.set_ytick
#
# # Shrink current axis by 20%
# box = ax.get_position()
# ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])
#
# handles, labels = ax.get_legend_handles_labels()
# ax.legend(handles[::-1], labels[::-1], loc='center left', \
#           bbox_to_anchor=(1, 0.5))
# # ax.legend()
#
# plt.savefig('emissions_breakdown.png', dpi=300)
