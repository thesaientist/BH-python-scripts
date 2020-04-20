# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 12:31:23 2019

@author: 212566876
"""

#import seaborn as sns
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np
from matplotlib.ticker import AutoMinorLocator

SMALL_SIZE = 8
MEDIUM_SIZE = 10
BIGGER_SIZE = 12

plt.rc('font', size=SMALL_SIZE)          # controls default text sizes
plt.rc('axes', titlesize=SMALL_SIZE)     # fontsize of the axes title
plt.rc('axes', labelsize=SMALL_SIZE)    # fontsize of the x and y labels
plt.rc('xtick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('ytick', labelsize=SMALL_SIZE)    # fontsize of the tick labels
plt.rc('legend', fontsize=SMALL_SIZE)    # legend fontsize
plt.rc('figure', titlesize=BIGGER_SIZE)  # fontsize of the figure title

# load workbook
# wb = openpyxl.load_workbook('LAR country data and plots.xlsx', data_only=True)
# ws = wb['Carbon Tax Lifting Cost Impact']

# load workbook
wb = openpyxl.load_workbook('../AAPG/AAPG_LatinAmerica_OPGEE.xlsm', data_only=True)
ws = wb['Results']

# read in needed columns for plto data
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
api = []            # api gravity of crude
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    api.append(ws.cell(row=34, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1

# convert lists to pandas dataframe
data = {'Country':countries, 'Field':fields, 'Production':production, \
        'API Gravity':api, 'Flaring (scf/bbl)':flaring_ratio, \
        'Exploration Combustion CI':exp_comb, \
        'Exploration VFF CI':exp_vff, \
        'Drilling Combustion CI':dnd_comb, \
        'Drilling VFF CI':dnd_vff, \
        'Production Combustion CI':prod_comb, \
        'Production VFF CI':prod_vff, \
        'Processing Combustion CI':proc_comb, \
        'Processing VFF CI':proc_vff, \
        'Maintenance Combustion CI':main_comb, \
        'Maintenance VFF CI':main_vff, \
        'Transportation CI':trans, \
        'Miscellaneous CI':misc, \
        'Offsite CI':offsite, \
        'Total CI':net_CI
    }
emissions = pd.DataFrame(data)
emissions['Total VFF CI'] = emissions['Exploration VFF CI'] + \
                            emissions['Drilling VFF CI'] + \
                            emissions['Production VFF CI'] + \
                            emissions['Processing VFF CI'] + \
                            emissions['Maintenance VFF CI']
emissions = emissions.sort_values(by='Total CI')

# get country averages
def wavg(group, avg_name, weight_name):
    """ http://stackoverflow.com/questions/10951341/pandas-dataframe-aggregate-function-using-multiple-columns
    In rare instance, we may not have weights, so just return the mean. Customize this if your business case
    should return otherwise.
    """
    d = group[avg_name]
    w = group[weight_name]
    try:
        return (d * w).sum() / w.sum()
    except ZeroDivisionError:
        return d.mean()

parameter_lst = list(emissions.columns)[3:]
country_data = {}
country_data['Production'] = emissions.groupby('Country')['Production'].sum()
for param in parameter_lst:
    country_data[param] = emissions.groupby('Country').apply(wavg, param, 'Production')

country_emissions = pd.DataFrame(data=country_data)     # country averages (g/MJ units)
country_emissions = country_emissions.sort_values(by='Total CI')

#################################################################
# convert gCO2e/MJ to tCO2e/bbl and come up with country average
#################################################################

# get fuel spec table imported into dataframe
fsSheet = wb['Fuel Specs']
api = np.zeros((42,))
lhv = np.zeros((42,))
for i in range(api.shape[0]):
    api[i] = fsSheet.cell(row=i+24, column=2).value
    lhv[i] = fsSheet.cell(row=i+24, column=12).value

# convert CIs into tCO2e/bbl
emissions['LHV'] = np.interp(emissions['API Gravity'], api, lhv)
emissions_bbl = emissions.copy()
for param in parameter_lst[2:]:
    emissions_bbl[param] = emissions[param] * emissions['LHV'] / 1e6

# get country averages
country_data = {}
country_data['Production'] = emissions_bbl.groupby('Country')['Production'].sum()
for param in parameter_lst:
    country_data[param] = emissions_bbl.groupby('Country').apply(wavg, param, 'Production')

country_emissions_bbl = pd.DataFrame(data=country_data)     # country averages (te/bbl units)
country_emissions_bbl = country_emissions_bbl.sort_values(by='Total CI')

#########################################################
# generate emissions breakdown plot for each country (g/MJ)
#########################################################

# separate United States from the rest (as it's used for benchmarking)
old_country_emissions = country_emissions.copy()
country_emissions = country_emissions.append(pd.Series(name='Blank 1'))
usa = old_country_emissions.loc['United States', :]
country_emissions = country_emissions.drop(index='United States')
country_emissions = country_emissions.append(usa)
country_emissions = country_emissions.append(pd.Series(name='Blank 2'))


N = len(country_emissions.index)
ind = np.arange(N)
expCI = country_emissions['Exploration Combustion CI']
dndCI = country_emissions['Drilling Combustion CI']
prodCI = country_emissions['Production Combustion CI']
procCI = country_emissions['Processing Combustion CI']
mainCI = country_emissions['Maintenance Combustion CI']
vffCI = country_emissions['Total VFF CI']
miscCI = country_emissions['Miscellaneous CI']
transCI = country_emissions['Transportation CI']
offsiteCI = country_emissions['Offsite CI']
totalCI = country_emissions['Total CI']

countryLabels = country_emissions.index.values.tolist()
countryLabels[countryLabels.index('Blank 1')] = ''
countryLabels[countryLabels.index('Blank 2')] = ''
countryLabels[countryLabels.index('United States')] = 'USA\nUnconv.'
countryLabels[countryLabels.index('Colombia')] = 'COL'
countryLabels[countryLabels.index('Brazil')] = 'BRA'
countryLabels[countryLabels.index('Argentina')] = 'ARG'
countryLabels[countryLabels.index('Mexico')] = 'MEX'
countryLabels[countryLabels.index('Ecuador')] = 'ECU'
countryLabels[countryLabels.index('Venezuela')] = 'VEN'

fig, ax = plt.subplots()
width = 0.75
exp_bar = ax.bar(ind, expCI, width, label='Exploration', \
                 color=(45/255, 96/255, 179/255, 0.8))
dnd_bar = ax.bar(ind, dndCI, width, bottom=expCI, label='Drilling', \
                 color=(232/255, 159/255, 12/255, 0.8))
prod_bar = ax.bar(ind, prodCI, width, bottom=expCI + dndCI, label='Production', \
                  color=(7/255, 71/255, 77/255, 0.8))
proc_bar = ax.bar(ind, procCI, width, bottom=expCI + dndCI + prodCI, \
                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
main_bar = ax.bar(ind, mainCI, width, bottom=expCI + dndCI + prodCI + procCI, \
            label='Maintenance', color=(201/255, 195/255, 0, 0.8))
vff_bar = ax.bar(ind, vffCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI, label='VFF', color=(255/255, 0, 0, 0.8))
misc_bar = ax.bar(ind, miscCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
trans_bar = ax.bar(ind, transCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI, label='Transport', \
            color=(69/255,74/255,65/255,0.8))
offsite_bar = ax.bar(ind, offsiteCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI + transCI, label='Offsite', \
            color=(0,128/255,1,0.8))
totalCI_line = ax.plot(ind, totalCI, 'kX-', label='Total CI')
# around the world comparison
saudi_line = ax.hlines(4.6, ind[-3], ind[-1], colors='b', linestyles='dotted')
canada_line = ax.hlines(17.6, ind[-3], ind[-1], colors='Orange', linestyles='dotted')
algeria_line = ax.hlines(20.30, ind[-3], ind[-1], colors='r', linestyles='dotted')

# add text labels to the benchmark lines
ax.text(ind[-3], 5, 'Saudi Arabia', color='b')
ax.text(ind[-3], 18, 'Canada', color='Orange')
ax.text(ind[-3], 20.7, 'Algeria', color='r')
ax.text(ind[-3], 23, 'BENCHMARKS', fontweight='bold')
ax.set_ylim(0,25)


ax.set_title('Carbon Intensity Comparison')
ax.set_xticks(ind.tolist())
ax.set_xticklabels(countryLabels)
ax.set_ylabel('Carbon Intensity (gCO2e/MJ crude oil)')
ax.yaxis.set_minor_locator(AutoMinorLocator())

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

handles, labels = ax.get_legend_handles_labels()
handles = handles[1:] + [handles[0]]
labels = labels[1:] + [labels[0]]
ax.legend(handles[::-1], labels[::-1], loc='center left', \
          bbox_to_anchor=(1, 0.5))

fig.savefig('emissions_breakdown.png', dpi=300)

#########################################################
# generate emissions breakdown plot for each country (te/bbl)
#########################################################
N = len(country_emissions_bbl.index)
ind = np.arange(N)
expCI = country_emissions_bbl['Exploration Combustion CI']
dndCI = country_emissions_bbl['Drilling Combustion CI']
prodCI = country_emissions_bbl['Production Combustion CI']
procCI = country_emissions_bbl['Processing Combustion CI']
mainCI = country_emissions_bbl['Maintenance Combustion CI']
vffCI = country_emissions_bbl['Total VFF CI']
miscCI = country_emissions_bbl['Miscellaneous CI']
transCI = country_emissions_bbl['Transportation CI']
offsiteCI = country_emissions_bbl['Offsite CI']
totalCI = country_emissions_bbl['Total CI']

countryLabels = country_emissions.index.values.tolist()
countryLabels[countryLabels.index('United States')] = 'USA\nUnconv.'
countryLabels[countryLabels.index('Colombia')] = 'COL'
countryLabels[countryLabels.index('Brazil')] = 'BRA'
countryLabels[countryLabels.index('Argentina')] = 'ARG'
countryLabels[countryLabels.index('Mexico')] = 'MEX'
countryLabels[countryLabels.index('Ecuador')] = 'ECU'
countryLabels[countryLabels.index('Venezuela')] = 'VEN'

fig, ax = plt.subplots()
width = 0.75
exp_bar = ax.bar(ind, expCI, width, label='Exploration', \
                 color=(45/255, 96/255, 179/255, 0.8))
dnd_bar = ax.bar(ind, dndCI, width, bottom=expCI, label='Drilling', \
                 color=(232/255, 159/255, 12/255, 0.8))
prod_bar = ax.bar(ind, prodCI, width, bottom=expCI + dndCI, label='Production', \
                  color=(7/255, 71/255, 77/255, 0.8))
proc_bar = ax.bar(ind, procCI, width, bottom=expCI + dndCI + prodCI, \
                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
main_bar = ax.bar(ind, mainCI, width, bottom=expCI + dndCI + prodCI + procCI, \
            label='Maintenance', color=(201/255, 195/255, 0, 0.8))
vff_bar = ax.bar(ind, vffCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI, label='VFF', color=(255/255, 0, 0, 0.8))
misc_bar = ax.bar(ind, miscCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
trans_bar = ax.bar(ind, transCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI, label='Transport', \
            color=(69/255,74/255,65/255,0.8))
offsite_bar = ax.bar(ind, offsiteCI, width, bottom=expCI + dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI + transCI, label='Offsite', \
            color=(0,128/255,1,0.8))
totalCI_line = ax.plot(ind, totalCI, 'kX-', label='Total CI')


ax.set_title('Carbon Intensity Comparison')
ax.set_xticks(ind.tolist())
ax.set_xticklabels(countryLabels)
ax.set_ylabel('Carbon Intensity (tCO2e/bbl crude oil)')

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

handles, labels = ax.get_legend_handles_labels()
handles = handles[1:] + [handles[0]]
labels = labels[1:] + [labels[0]]
ax.legend(handles[::-1], labels[::-1], loc='center left', \
          bbox_to_anchor=(1, 0.5))

plt.savefig('emissions_breakdown_bbl.png', dpi=300)

#########################################################
# cost calculations and incremental cost of supply curve
#########################################################

# Add base cost information for these countries
base_cost = {'Brazil':80.23, 'Colombia':92.23, 'Argentina':89.28, 'Mexico':55.98, 'Ecuador':62.68, 'Venezuela':66.61}
country_total_prod = {'Brazil':2.620, 'Colombia':0.854, 'Argentina':0.480, 'Mexico':1.981, 'Ecuador':0.531, 'Venezuela':2.010, 'United States':12.3}
country_emissions_bbl['Base Cost'] = country_emissions_bbl.index.map(base_cost)
country_emissions_bbl['Total Prod MMBOPD'] = country_emissions_bbl.index.map(country_total_prod)
country_emissions_bbl['Production'] = country_emissions_bbl['Production']/1e6   # convert sum of fields production for each country into mmbopd

# Add Bolivia to country list
#country_emissions_bbl.loc['Bolivia']['Base Cost'] = 78.21
#country_emissions_bbl.loc['Bolivia']['Total CI'] = 9.2*5818.2/1e6
#country_emissions_bbl.loc['Bolivia']['Total Prod MMBOPD'] = 0.05933
bolivia = pd.Series({'Base Cost':78.21, 'Total CI':9.2*5818.2/1e6, 'Total Prod MMBOPD':0.05933}, name='Bolivia')
country_emissions_bbl = country_emissions_bbl.append(bolivia)

# Calculate S1 and S2 costs and subsequently as percentages of base cost
country_emissions_bbl['S1 Cost'] = 50 * country_emissions_bbl['Total CI']
country_emissions_bbl['S2 Cost'] = 150 * country_emissions_bbl['Total CI']
country_emissions_bbl['S1 Percent'] = country_emissions_bbl['S1 Cost']/country_emissions_bbl['Base Cost']*100
country_emissions_bbl['S2 Percent'] = (country_emissions_bbl['S2 Cost']-country_emissions_bbl['S1 Cost'])/country_emissions_bbl['Base Cost']*100

# Sort and calculate cumulative production column for use in plot
#country_emissions_bbl = country_emissions_bbl.sort_values(by='Base Cost')
country_emissions_bbl = country_emissions_bbl.sort_values(by='S1 Percent')  # sort by S1 percent ($50/tCO2e carbon price incremental cost of supply)
#country_emissions_bbl = country_emissions_bbl.sort_values(by='S2 Percent')  # sort by S2 percent ($50/tCO2e carbon price incremental cost of supply)
country_emissions_bbl['Cumulative Production'] = country_emissions_bbl['Total Prod MMBOPD'].cumsum()

left = country_emissions_bbl.iloc[:-1]['Cumulative Production'].tolist()     # exclude last element in cumulative production (for use in plot)
left = [0.0] + left                                                         # include 0 at beginning of list for use in plot x-axis

# make lists for use in plot as bases for stacking columns
countries = country_emissions_bbl.index.values
production = country_emissions_bbl['Total Prod MMBOPD'].tolist()
# base_cost = country_emissions_bbl['Base Cost'].tolist()
s1_percent = country_emissions_bbl['S1 Percent'].tolist()
s2_percent = country_emissions_bbl['S2 Percent'].tolist()

### PLOT ###
fig, ax = plt.subplots()

#baseCol = ax.bar(left, base_cost, width = production,
#                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
#s1Col = ax.bar(left, s1_cost, width = production, bottom=base_cost,
#                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
s1Col = ax.bar(left, s1_percent, width = production,
                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
s2Col = ax.bar(left, s2_percent, width = production, bottom=s1_percent,
                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
#usUnConv = ax.hlines(us_unconv_cost, 0, cum_prod[-1], colors='maroon', linestyles='dashed', label='U.S. Unconventionals')

ax.set_ylim(0,22)
ax.set_xlabel('Total Crude Oil Production \n million bbl/d')
ax.set_ylabel('Incremental cost \n (%)')
ax.set_title('Carbon Pricing Impact on Costs in Latin America')
#ax.legend((baseCol[0], s1Col[0], s2Col[0], usUnConv), ('Base', '\$50/tCO2e', '\$150/tCO2e', 'U.S. Unconventionals'))
#ax.legend((baseCol[0], s1Col[0], s2Col[0]), ('Base', '\$50/tCO2e', '\$150/tCO2e'))
ax.legend((s1Col[0], s2Col[0]), ('\$50/tCO2e', '\$150/tCO2e'), loc='upper left')
for i in range(7):
    if i == 0:
        ax.text((0 + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.1, 6.8, countries[i], rotation=90)
    elif i==1:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.1, 6.8, countries[i], rotation=90)
    elif i==2:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.45, 1, countries[i], rotation=0)
    elif i==3:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.25, 15, countries[i], rotation=90)
    elif i==4:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.1, 3, countries[i], rotation=90)
    elif i==5:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.45, 1, countries[i], rotation=0)
    elif i==6:
        ax.text((country_emissions_bbl.iloc[i-1]['Cumulative Production'] + country_emissions_bbl.iloc[i]['Cumulative Production'])/2-0.45, 1, countries[i], rotation=0)
    # peru
#    elif i==3:
#        ax.text((cum_prod[i] + cum_prod[i+1])/2-0.35, 100, countries[i], rotation=90)
plt.savefig('cost_curve.png', bbox_inches = 'tight', dpi=300)

#######
# MISC
#######

# calculate total annual emissions from VFF for each country (million tCO2e)
annual_vff = (country_emissions_bbl['Total VFF CI']*country_emissions_bbl['Total Prod MMBOPD']*365).sort_values(ascending=True)
annual_vff = annual_vff.drop(index=['United States', 'Bolivia'])
fig, ax = plt.subplots()
ax = annual_vff.plot(kind='barh', color='blue')
ax.set_xlabel('Annual VFF Emissions (million tCO2e)')
fig.savefig('annual_vff.png', dpi=300)
