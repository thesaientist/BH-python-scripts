# -*- coding: utf-8 -*-
"""
Created on Mon Jun  3 12:31:23 2019

@author: 212566876
"""

#import seaborn as sns
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd

# define any required constants
us_unconv_cost = 50      # $/bbl lifting cost for US Unconventional Production (2016 number, Rystad, as quoted in WSJ)

# load workbook
wb = openpyxl.load_workbook('LAM country data and plots.xlsx', data_only=True)
ws = wb['Carbon Tax Lifting Cost Impact']

# read in needed columns for plto data
countries = []
production = []     # country production rates (MMbbl/d), used for width of columns
base_cost = []      # base lifting cost ($/bbl)
s1_cost = []        # scenario 1 lifting cost (increment over base)
s2_cost = []        # scenario 2 lifting cost (increment over S1)

# collect country data
for i in range(3,11):
    # countries.append(ws.cell(row=i, column=1).value)      # Country shown
    countries.append("Country " + str(i-2))                 # Anonymous countries
    production.append(ws.cell(row=i, column=2).value)
    #cum_prod.append(ws.cell(row=i, column=3).value)
    base_cost.append(ws.cell(row=i, column=4).value)
    s1_cost.append(ws.cell(row=i, column=17).value)
    s2_cost.append(ws.cell(row=i, column=18).value)

# convert lists to pandas dataframe
data = {'Country':countries, 'Production':production, 'Base Cost':base_cost, 'S1 Cost':s1_cost, 'S2 Cost':s2_cost}
lam = pd.DataFrame(data)
lam = lam.set_index('Country')
# lam = lam.drop(index='Peru')
lam = lam.drop(index='Country 6')       # Anonymous
#lam = lam.drop(index='Bolivia')


# Calculate S1 and S2 costs as percentages of base cost
lam['S1 Percent'] = lam['S1 Cost']/lam['Base Cost']*100
lam['S2 Percent'] = lam['S2 Cost']/lam['Base Cost']*100

# Sort by base cost and calculate cumulative production column for use in plot
#lam = lam.sort_values(by='Base Cost')
lam = lam.sort_values(by='S1 Percent')  # sort by S1 percent ($50/tCO2e carbon price incremental cost of supply)
#lam = lam.sort_values(by='S2 Percent')  # sort by S2 percent ($50/tCO2e carbon price incremental cost of supply)
lam['Cumulative Production'] = lam['Production'].cumsum()
#lowest_cost = base_cost[0]
#for i in range(len(base_cost)):
#    base_cost[i] -= lowest_cost    # convert to relative cost

left = lam.iloc[:-1]['Cumulative Production'].tolist()     # exclude last element in cumulative production (for use in plot)
left = [0.0] + left                                          # include 0 at beginning of list for use in plot x-axis

# make lists for use in plot as bases for stacking columns
countries = lam.index.values
production = lam['Production'].tolist()
base_cost = lam['Base Cost'].tolist()
s1_percent = lam['S1 Percent'].tolist()
s2_percent = lam['S2 Percent'].tolist()

#s1_cum_cost = [base_cost[i] + s1_cost[i] for i in range(len(base_cost))]

### PLOT ###
fig, ax = plt.subplots()

#baseCol = ax.bar(left, base_cost, width = production,
#                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
#s1Col = ax.bar(left, s1_cost, width = production, bottom=base_cost,
#                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
s1Col = ax.bar(left, s1_percent, width = production,
                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
s2Col = ax.bar(left, s2_percent, width = production, bottom=s1_percent,
                  alpha = 0.6, align='edge', edgecolor = 'k', linewidth = 1)
#usUnConv = ax.hlines(us_unconv_cost, 0, cum_prod[-1], colors='maroon', linestyles='dashed', label='U.S. Unconventionals')

ax.set_ylim(0,22)
ax.set_xlabel('Total Crude Oil Production \n million bbl/d')
ax.set_ylabel('Incremental cost \n (%)')
ax.set_title('Carbon Pricing Impact on Crude Supply Costs')
#ax.legend((baseCol[0], s1Col[0], s2Col[0], usUnConv), ('Base', '\$50/tCO2e', '\$150/tCO2e', 'U.S. Unconventionals'))
#ax.legend((baseCol[0], s1Col[0], s2Col[0]), ('Base', '\$50/tCO2e', '\$150/tCO2e'))
ax.legend((s1Col[0], s2Col[0]), ('\$50/tCO2e', '\$150/tCO2e'), loc='upper left')
for i in range(7):
    if i == 0:
        ax.text((0 + lam.iloc[i]['Cumulative Production'])/2-0.45, 1, countries[i], rotation=0)
    elif i==1:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.1, 3, countries[i], rotation=90)
    elif i==2:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.1, 4, countries[i], rotation=90)
    elif i==3:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.25, 12, countries[i], rotation=90)
    elif i==4:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.10, 6.5, countries[i], rotation=90)
    elif i==5:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.45, 1, countries[i], rotation=0)
    elif i==6:
        ax.text((lam.iloc[i-1]['Cumulative Production'] + lam.iloc[i]['Cumulative Production'])/2-0.65, 1, countries[i], rotation=0)
    # peru
#    elif i==3:
#        ax.text((cum_prod[i] + cum_prod[i+1])/2-0.35, 100, countries[i], rotation=90)

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0 + box.width*0.1, box.width, box.height])
plt.savefig('cost_curve.png', bbox_inches = 'tight', dpi=300)
