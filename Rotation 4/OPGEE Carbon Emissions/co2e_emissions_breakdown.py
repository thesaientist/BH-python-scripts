# -*- coding: utf-8 -*-
"""
Created on Thu Aug  8 10:45:38 2019

@author: 212566876
"""

#import seaborn as sns
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import numpy as np

# load workbook
wb = openpyxl.load_workbook('./Model/OPGEE_v2.0 - Simple Onshore Field - Sensitivities.xlsm', data_only=True)
ws = wb['Results']
ws_prod = wb['Production Profile']

# read in data for plotting field profile
year = []
oil_rate = []
gor = []
water_cut = []

# collect data on field profile
col = 3
while col<23:
    year.append(ws_prod.cell(row=3, column=col).value)
    oil_rate.append(ws_prod.cell(row=4, column=col).value)
    gor.append(ws_prod.cell(row=15, column=col).value)
    water_cut.append(ws_prod.cell(row=13, column=col).value)
    col += 1

# plot field profile
fig, ax1 = plt.subplots()

#color = 'tab:red'
ax1.set_xlabel('Year')
ax1.set_ylabel('Oil Production (bbl/d)\nGOR (scf/bbl)')
ax1.set_title('Field Profile')
lns1 = ax1.plot(year, oil_rate, color=(0/255, 100/255, 0/255, 0.8), label='Oil Rate')
lns2 = ax1.plot(year, gor, 'v',color=(255/255, 0/255, 0/255, 0.8), label='GOR')
ax1.xaxis.set_ticks(np.arange(5, 20, 5))
#ax1.grid(True)
#ax1.tick_params(axis='y', labelcolor=color)

ax2 = ax1.twinx()  # instantiate a second axes that shares the same x-axis

#color = 'tab:blue'
ax2.set_ylabel('Water Cut')  # we already handled the x-label with ax1
lns3 = ax2.plot(year, water_cut, 's',color=(30/255, 144/255, 255/255, 0.8), label='Water Cut')
ax2.set_ylim(0,1)
#ax2.tick_params(axis='y', labelcolor=color)

# added these three lines to legend
lns = lns1+lns2+lns3
labs = [l.get_label() for l in lns]
box = ax1.get_position()
ax1.set_position([box.x0, box.y0, box.width * 0.8, box.height])
ax1.legend(lns, labs, loc='center left', \
          bbox_to_anchor=(1.15, 0.5))

#fig.suptitle('Field Profile')
fig.tight_layout()  # otherwise the right y-label is slightly clipped
fig.savefig('field_profile.png', dpi=300)


# read in needed columns for plot data for emissions
countries = []      # names of countries
fields = []         # names of fields
production = []     # field production rates (bbl/d)
flaring_ratio = []  # flaring-to-oil ratio (scf gas flared/bbl of crude produced)
# lists allocated for all CI below (g/MJ)
exp_comb = []
exp_vff = []
dnd_comb = []
dnd_vff = []
prod_comb = []
prod_vff = []
proc_comb = []
proc_vff = []
main_comb = []
main_vff = []
trans = []
misc = []
offsite = []
net_CI = []

# collect data until last existing column
col = 8
while True:
    countries.append(ws.cell(row=20, column=col).value)
    fields.append(ws.cell(row=21, column=col).value)
    production.append(ws.cell(row=24, column=col).value)
    flaring_ratio.append(ws.cell(row=86, column=col).value)
    exp_comb.append(ws.cell(row=132, column=col).value)
    exp_vff.append(ws.cell(row=133, column=col).value)
    dnd_comb.append(ws.cell(row=138, column=col).value)
    dnd_vff.append(ws.cell(row=139, column=col).value)
    prod_comb.append(ws.cell(row=144, column=col).value)
    prod_vff.append(ws.cell(row=145, column=col).value)
    proc_comb.append(ws.cell(row=150, column=col).value)
    proc_vff.append(ws.cell(row=151, column=col).value)
    main_comb.append(ws.cell(row=156, column=col).value)
    main_vff.append(ws.cell(row=157, column=col).value)
    trans.append(ws.cell(row=167, column=col).value)
    misc.append(ws.cell(row=170, column=col).value)
    offsite.append(ws.cell(row=172, column=col).value)
    net_CI.append(ws.cell(row=179, column=col).value)

    nextField = ws.cell(row=21, column=col+1).value
    if not nextField:
        break
    else:
        col += 1

# convert lists to pandas dataframe
data = {'Country':countries, 'Field':fields, 'Production':production, \
        'Flaring (scf/bbl)':flaring_ratio, \
        'Exploration Combustion CI':exp_comb, \
        'Exploration VFF CI':exp_vff, \
        'Drilling Combustion CI':dnd_comb, \
        'Drilling VFF CI':dnd_vff, \
        'Production Combustion CI':prod_comb, \
        'Production VFF CI':prod_vff, \
        'Processing Combustion CI':proc_comb, \
        'Processing VFF CI':proc_vff, \
        'Maintenance Combustion CI':main_comb, \
        'Maintenance VFF CI':main_vff, \
        'Transportation CI':trans, \
        'Miscellaneous CI':misc, \
        'Offsite CI':offsite, \
        'Total CI':net_CI
    }
emissions = pd.DataFrame(data)
emissions['Total VFF CI'] = emissions['Exploration VFF CI'] + \
                            emissions['Drilling VFF CI'] + \
                            emissions['Production VFF CI'] + \
                            emissions['Processing VFF CI'] + \
                            emissions['Maintenance VFF CI']
emissions = emissions.sort_values(by='Total CI')

# generate emissions breakdown plot for each field
N = len(emissions.index)
ind = np.arange(N)
dndCI = emissions['Drilling Combustion CI']
prodCI = emissions['Production Combustion CI']
procCI = emissions['Processing Combustion CI']
mainCI = emissions['Maintenance Combustion CI']
vffCI = emissions['Total VFF CI']
miscCI = emissions['Miscellaneous CI']
transCI = emissions['Transportation CI']
offsiteCI = emissions['Offsite CI']

fieldLabels = emissions['Field'].tolist()
# legendLabels = ['Drilling', 'Production', 'Processing', 'Maintenance', 'VFF', \
#                 'Misc.', 'Transport']

fig, ax = plt.subplots()
width = 0.75
# color=(45/255, 96/255, 179/255, 0.8)  # color for exploration bar if ever added
dnd_bar = ax.bar(ind, dndCI, width, label='Drilling', \
                 color=(232/255, 159/255, 12/255, 0.8))
prod_bar = ax.bar(ind, prodCI, width, bottom=dndCI, label='Production', \
                  color=(7/255, 71/255, 77/255, 0.8))
proc_bar = ax.bar(ind, procCI, width, bottom=dndCI + prodCI, \
                  label='Processing', color=(247/255, 208/255, 12/255, 0.8))
main_bar = ax.bar(ind, mainCI, width, bottom=dndCI + prodCI + procCI, \
            label='Maintenance', color=(201/255, 195/255, 0, 0.8))
vff_bar = ax.bar(ind, vffCI, width, bottom=dndCI + prodCI + procCI + \
            mainCI, label='VFF', color=(225/255, 10/255, 0, 0.8))
misc_bar = ax.bar(ind, miscCI, width, bottom=dndCI + prodCI + procCI + \
            mainCI + vffCI, label='Misc.', color=(164/255,225/255,237/255,0.8))
trans_bar = ax.bar(ind, transCI, width, bottom=dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI, label='Transport', \
            color=(69/255,74/255,65/255,0.8))
offsite_bar = ax.bar(ind, offsiteCI, width, bottom=dndCI + prodCI + procCI + \
            mainCI + vffCI + miscCI + transCI, label='Offsite', \
            color=(0,128/255,1,0.8))

ax.set_title('GHG Emissions Intensity Comparison')
ax.set_xticks(ind.tolist())
ax.set_xticklabels(fieldLabels)
ax.set_ylabel('GHG Intensity (gCO2e/MJ crude oil)')
# ax.set_ytick

# Shrink current axis by 20%
box = ax.get_position()
ax.set_position([box.x0, box.y0, box.width * 0.8, box.height])

handles, labels = ax.get_legend_handles_labels()
ax.legend(handles[::-1], labels[::-1], loc='center left', \
          bbox_to_anchor=(1, 0.5))
# ax.legend()

plt.savefig('emissions_breakdown.png', dpi=300)
