################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2018
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
# import numpy as np
# import matplotlib.pyplot as plt
# import pandas as pd
# import Frac_ML_Library as fracmlLib

# path_to_folder = 'E:/Working/0 Submission CF approved annex5/Aprobadas y ajustadas por eco 19-01-2019/adjusted by eco/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/adjusted by eco/'
path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Approved by GCA adjusted by GCA(SAI)/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/test approved/'
# cashFlowFiles = []

# Get column numbers
colNum_C = xlLib.col2num('C')
colNum_D = xlLib.col2num('D')
colNum_G = xlLib.col2num('G')
colNum_H = xlLib.col2num('H')
colNum_J = xlLib.col2num('J')
colNum_K = xlLib.col2num('K')
colNum_L = xlLib.col2num('L')
colNum_O = xlLib.col2num('O')
colNum_P = xlLib.col2num('P')
colNum_R = xlLib.col2num('R')
colNum_S = xlLib.col2num('S')
colNum_T = xlLib.col2num('T')
colNum_W = xlLib.col2num('W')
colNum_AA = xlLib.col2num('AA')
colNum_AB = xlLib.col2num('AB')
colNum_AC = xlLib.col2num('AC')
colNum_AD = xlLib.col2num('AD')
colNum_AE = xlLib.col2num('AE')
colNum_AF = xlLib.col2num('AF')
colNum_AG = xlLib.col2num('AG')


# for loop to get all the workbook objects to modify, make the modifications needed and save

for i in os.listdir(path_to_folder):
    if i.endswith('.xlsx'):
        startTime = time.time()                                             # start timer for loop duration (DEBUG)
        cashFlowFilePath = os.path.join(path_to_folder, i)                  # get cash flow workbook file path
        wb = openpyxl.load_workbook(cashFlowFilePath)                       # load workbook

        # # modification 1
        # ws_PS = wb['CF PS']
        # ws_PS['AF51'] = '=SUM(AF13:AF49)'
        # ws_PS['AG51'] = '=SUM(AG13:AG49)'
        
        # SHEET CF PDP
        ws_PDP = wb['CF PDP']
        for rowNum in range(13,50):
            ws_PDP.cell(row=rowNum, column=colNum_R).value = "='Data CF PDP'!AP" + str(rowNum+23) +"/1000000"
        
        # SHEET CF PNP
        ws_PNP = wb['CF PNP']
        for rowNum in range(13,50):
            ws_PNP.cell(row=rowNum, column=colNum_C).value = "='Data CF PNP'!S" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_D).value = "='Data CF PNP'!T" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_G).value = "=('Data CF PNP'!W" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!X" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!Y" + str(rowNum+23) + ")/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_H).value = "='Data CF PNP'!AC" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_J).value = "='Data CF PNP'!AE" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_K).value = "='Data CF PNP'!AG" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_L).value = "='Data CF PNP'!AH" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_O).value = "=('Data CF PNP'!AK" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!AL" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!AM" + str(rowNum+23) + ")/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_P).value = "='Data CF PNP'!AN" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_R).value = "='Data CF PNP'!AP" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_S).value = "='Data CF PNP'!AQ" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_T).value = "='Data CF PNP'!AR" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_W).value = "=('Data CF PNP'!AU" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!AV" + str(rowNum+23) + \
                                                              "+'Data CF PNP'!AW" + str(rowNum+23) + ")/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AA).value = "='Data CF PNP'!BF" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AB).value = "='Data CF PNP'!BG" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AC).value = "='Data CF PNP'!CG" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AD).value = "=('Data CF PNP'!CU" + str(rowNum+23) + \
                                                              "*('Data CF PNP'!CV" + str(rowNum+23) + \
                                                              "/100))/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AE).value = "='Data CF PNP'!CW" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AF).value = "='Data CF PNP'!CY" + str(rowNum+23) +"/1000000"
            ws_PNP.cell(row=rowNum, column=colNum_AG).value = "='Data CF PNP'!CZ" + str(rowNum+23) +"/1000000"

        # SHEET CF PND
        ws_PND = wb['CF PND']
        for rowNum in range(13,50):
            ws_PND.cell(row=rowNum, column=colNum_C).value = "='Data CF PND'!S" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_D).value = "='Data CF PND'!T" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_G).value = "=('Data CF PND'!W" + str(rowNum+23) + \
                                                              "+'Data CF PND'!X" + str(rowNum+23) + \
                                                              "+'Data CF PND'!Y" + str(rowNum+23) + ")/1000000"
            ws_PND.cell(row=rowNum, column=colNum_H).value = "='Data CF PND'!AC" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_J).value = "='Data CF PND'!AE" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_K).value = "='Data CF PND'!AG" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_L).value = "='Data CF PND'!AH" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_O).value = "=('Data CF PND'!AK" + str(rowNum+23) + \
                                                              "+'Data CF PND'!AL" + str(rowNum+23) + \
                                                              "+'Data CF PND'!AM" + str(rowNum+23) + ")/1000000"
            ws_PND.cell(row=rowNum, column=colNum_P).value = "='Data CF PND'!AN" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_R).value = "='Data CF PND'!AP" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_S).value = "='Data CF PND'!AQ" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_T).value = "='Data CF PND'!AR" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_W).value = "=('Data CF PND'!AU" + str(rowNum+23) + \
                                                              "+'Data CF PND'!AV" + str(rowNum+23) + \
                                                              "+'Data CF PND'!AW" + str(rowNum+23) + ")/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AA).value = "='Data CF PND'!BF" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AB).value = "='Data CF PND'!BG" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AC).value = "='Data CF PND'!CG" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AD).value = "=('Data CF PND'!CU" + str(rowNum+23) + \
                                                              "*('Data CF PND'!CV" + str(rowNum+23) + \
                                                              "/100))/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AE).value = "='Data CF PND'!CW" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AF).value = "='Data CF PND'!CY" + str(rowNum+23) +"/1000000"
            ws_PND.cell(row=rowNum, column=colNum_AG).value = "='Data CF PND'!CZ" + str(rowNum+23) +"/1000000"

        # SHEET CF PRB
        ws_PRB = wb['CF PRB']
        for rowNum in range(13,50):
            ws_PRB.cell(row=rowNum, column=colNum_C).value = "='Data CF PRB'!S" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_D).value = "='Data CF PRB'!T" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_G).value = "=('Data CF PRB'!W" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!X" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!Y" + str(rowNum+23) + ")/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_H).value = "='Data CF PRB'!AC" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_J).value = "='Data CF PRB'!AE" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_K).value = "='Data CF PRB'!AG" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_L).value = "='Data CF PRB'!AH" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_O).value = "=('Data CF PRB'!AK" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!AL" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!AM" + str(rowNum+23) + ")/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_P).value = "='Data CF PRB'!AN" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_R).value = "='Data CF PRB'!AP" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_S).value = "='Data CF PRB'!AQ" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_T).value = "='Data CF PRB'!AR" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_W).value = "=('Data CF PRB'!AU" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!AV" + str(rowNum+23) + \
                                                              "+'Data CF PRB'!AW" + str(rowNum+23) + ")/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AA).value = "='Data CF PRB'!BF" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AB).value = "='Data CF PRB'!BG" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AC).value = "='Data CF PRB'!CG" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AD).value = "=('Data CF PRB'!CU" + str(rowNum+23) + \
                                                              "*('Data CF PRB'!CV" + str(rowNum+23) + \
                                                              "/100))/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AE).value = "='Data CF PRB'!CW" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AF).value = "='Data CF PRB'!CY" + str(rowNum+23) +"/1000000"
            ws_PRB.cell(row=rowNum, column=colNum_AG).value = "='Data CF PRB'!CZ" + str(rowNum+23) +"/1000000"
        
        # SHEET CF PS
        ws_PS = wb['CF PS']
        for rowNum in range(13,50):
            ws_PS.cell(row=rowNum, column=colNum_C).value = "='Data CF PS'!S" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_D).value = "='Data CF PS'!T" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_G).value = "=('Data CF PS'!W" + str(rowNum+23) + \
                                                              "+'Data CF PS'!X" + str(rowNum+23) + \
                                                              "+'Data CF PS'!Y" + str(rowNum+23) + ")/1000000"
            ws_PS.cell(row=rowNum, column=colNum_H).value = "='Data CF PS'!AC" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_J).value = "='Data CF PS'!AE" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_K).value = "='Data CF PS'!AG" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_L).value = "='Data CF PS'!AH" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_O).value = "=('Data CF PS'!AK" + str(rowNum+23) + \
                                                              "+'Data CF PS'!AL" + str(rowNum+23) + \
                                                              "+'Data CF PS'!AM" + str(rowNum+23) + ")/1000000"
            ws_PS.cell(row=rowNum, column=colNum_P).value = "='Data CF PS'!AN" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_R).value = "='Data CF PS'!AP" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_S).value = "='Data CF PS'!AQ" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_T).value = "='Data CF PS'!AR" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_W).value = "=('Data CF PS'!AU" + str(rowNum+23) + \
                                                              "+'Data CF PS'!AV" + str(rowNum+23) + \
                                                              "+'Data CF PS'!AW" + str(rowNum+23) + ")/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AA).value = "='Data CF PS'!BF" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AB).value = "='Data CF PS'!BG" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AC).value = "='Data CF PS'!CG" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AD).value = "=('Data CF PS'!CU" + str(rowNum+23) + \
                                                              "*('Data CF PS'!CV" + str(rowNum+23) + \
                                                              "/100))/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AE).value = "='Data CF PS'!CW" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AF).value = "='Data CF PS'!CY" + str(rowNum+23) +"/1000000"
            ws_PS.cell(row=rowNum, column=colNum_AG).value = "='Data CF PS'!CZ" + str(rowNum+23) +"/1000000"
        
        wb.save(cashFlowFilePath)                                           # save workbook

        # helpful message after saving workbook
        print('Workbook ' + os.path.basename(cashFlowFilePath) + ' is saved.')
        
        stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
        print('Time elapsed: ' + str(stopTime-startTime))
        # ask for confirmation before proceeding to next file (DEBUG)
        # wait = input('Press Enter to continue to the next workbook...')



print('Successfully done editing!')
