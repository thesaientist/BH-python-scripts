################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2018
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
import numpy as np

path_to_source = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Annex 4 Jesus/Eco CF/'
path_to_annex3 = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Annex 4 Jesus/annex 3/'
path_to_target = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Annex 4 Jesus/CRColas/'

# Get column numbers
# source docs
colNum_AQ = xlLib.col2num('AQ')     # net Oil (BBL)
colNum_AS = xlLib.col2num('AS')     # net gas (MCF)
colNum_AU = xlLib.col2num('AU')     # net propane (BBL)
colNum_AV = xlLib.col2num('AV')     # net butane (BBL)
colNum_AW = xlLib.col2num('AW')     # net gasoline (BBL)

# -----------------------ANNEX 5------------------------------------------------
# source defintions & allocations
cashFlowList = [file for file in os.listdir(path_to_source) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_source, file))]
cashFlowList.sort()     # sort alphabetically
numFields = len(cashFlowList)
netOil = np.zeros((61,10,numFields))
netGas = np.zeros((61,10,numFields))
netProp = np.zeros((61,10,numFields))
netBut = np.zeros((61,10,numFields))
netGasol = np.zeros((61,10,numFields))

# for loop to get all the data from annex 5
for i, file in enumerate(cashFlowList):
    startTime = time.time()                                             # start timer for loop duration (DEBUG)
    cashFlowFilePath = os.path.join(path_to_source, file)                  # get cash flow workbook file path
    wb = openpyxl.load_workbook(cashFlowFilePath, data_only=True)       # load workbook (read-only mode, faster)
    ws_PDP = wb['Data CF-PDP']
    ws_PNP = wb['Data CF-PNP']
    ws_PND = wb['Data CF-PND']
    ws_PRB = wb['Data CF-PRB']
    ws_PS = wb['Data CF-PS']
    wks = [ws_PDP, ws_PNP, ws_PND, ws_PRB, ws_PS]

    # gather data from all the sheets in the cash flow
    for k in range(len(wks)):
        for j in range(12, 73):
            netOil[j-12, k, i] = wks[k].cell(row=j, column=colNum_AQ).value
            netGas[j-12, k, i] = wks[k].cell(row=j, column=colNum_AS).value
            netProp[j-12, k, i] = wks[k].cell(row=j, column=colNum_AU).value
            netBut[j-12, k, i] = wks[k].cell(row=j, column=colNum_AV).value
            netGasol[j-12, k, i] = wks[k].cell(row=j, column=colNum_AW).value

    # helpful message after saving workbook
    print('Workbook ' + os.path.basename(cashFlowFilePath) + ' from Annex 5 is extracted.')
    stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
    print('Time elapsed: ' + str(stopTime-startTime))
    # ask for confirmation before proceeding to next file (DEBUG)
    # wait = input('Press Enter to continue to the next workbook...')

# save np arrays to disk to allow reloading them without having to run openpyxl again
np.save('./npy annex 5/netOil.npy', netOil)
np.save('./npy annex 5/netGas.npy', netGas)
np.save('./npy annex 5/netProp.npy', netProp)
np.save('./npy annex 5/netBut.npy', netBut)
np.save('./npy annex 5/netGasol.npy', netGasol)

print('Numpy databases for annex 5 have been saved.')

# -----------------------ANNEX 3 added data-------------------------------------
netOil = np.load('./npy annex 5/netOil.npy')
netGas = np.load('./npy annex 5/netGas.npy')
netProp = np.load('./npy annex 5/netProp.npy')
netBut = np.load('./npy annex 5/netBut.npy')
netGasol = np.load('./npy annex 5/netGasol.npy')
# source defintions & allocations
annex3List = [file for file in os.listdir(path_to_annex3) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_annex3, file))]
annex3List.sort()     # sort alphabetically
numFields = len(annex3List)
# for loop to get all the data from annex 3 (prb breakdown and ps breakdown)
for i, file in enumerate(annex3List):
    startTime = time.time()                                             # start timer for loop duration (DEBUG)
    annex3FilePath = os.path.join(path_to_annex3, file)                  # get cash flow workbook file path
    wb = openpyxl.load_workbook(annex3FilePath, data_only=True)       # load workbook (read-only mode, faster)
    ws_consol = wb['Consolidated Forecast']

    # gather data from all the sheets in the cash flow
    for j in range(6, 67):
        netOil[j-6, k, i] = wks[k].cell(row=j, column=colNum_AQ).value
        netGas[j-6, k, i] = wks[k].cell(row=j, column=colNum_AS).value
        netProp[j-6, k, i] = wks[k].cell(row=j, column=colNum_AU).value
        netBut[j-6, k, i] = wks[k].cell(row=j, column=colNum_AV).value
        netGasol[j-6, k, i] = wks[k].cell(row=j, column=colNum_AW).value

    # helpful message after saving workbook
    print('Workbook ' + os.path.basename(annex3FilePath) + ' from Annex 3 is extracted.')
    stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
    print('Time elapsed: ' + str(stopTime-startTime))
    # ask for confirmation before proceeding to next file (DEBUG)
    # wait = input('Press Enter to continue to the next workbook...')

# save np arrays to disk to allow reloading them without having to run openpyxl again
np.save('./npy annex 3 and 5/netOil.npy', netOil)
np.save('./npy annex 3 and 5/netGas.npy', netGas)
np.save('./npy annex 3 and 5/netProp.npy', netProp)
np.save('./npy annex 3 and 5/netBut.npy', netBut)
np.save('./npy annex 3 and 5/netGasol.npy', netGasol)

print('Numpy databases for annex 3 have been saved.')

#-----------------------------ADD to CRCOLAS------------------------------------
netOil = np.load('./npy annex 3 and 5/netOil.npy')
netGas = np.load('./npy annex 3 and 5/netGas.npy')
netProp = np.load('./npy annex 3 and 5/netProp.npy')
netBut = np.load('./npy annex 3 and 5/netBut.npy')
netGasol = np.load('./npy annex 3 and 5/netGasol.npy')

# target defintions & allocations
targetList = [file for file in os.listdir(path_to_target) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_target, file))]
targetList.sort()     # sort alphabetically
numFields = len(targetList)
# for loop to get add the data to target
for i, file in enumerate(targetList):
    startTime = time.time()                                             # start timer for loop duration (DEBUG)
    targetFilePath = os.path.join(path_to_target, file)                  # get cash flow workbook file path
    wb = openpyxl.load_workbook(targetFilePath)                         # load workbook
    ws_consol = wb['Consolidated Forecast']

    # gather data from all the sheets in the cash flow
    for j in range(6, 67):
        netOil[j-6, k, i] = wks[k].cell(row=j, column=colNum_AQ).value
        netGas[j-6, k, i] = wks[k].cell(row=j, column=colNum_AS).value
        netProp[j-6, k, i] = wks[k].cell(row=j, column=colNum_AU).value
        netBut[j-6, k, i] = wks[k].cell(row=j, column=colNum_AV).value
        netGasol[j-6, k, i] = wks[k].cell(row=j, column=colNum_AW).value

    # helpful message after saving workbook
    print('Workbook ' + os.path.basename(annex3FilePath) + ' modified.')
    stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
    print('Time elapsed: ' + str(stopTime-startTime))
    # ask for confirmation before proceeding to next file (DEBUG)
    # wait = input('Press Enter to continue to the next workbook...')
