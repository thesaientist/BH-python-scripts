################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2018
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
import numpy as np
# import matplotlib.pyplot as plt
# import pandas as pd

# path_to_folder = 'E:/Working/0 Submission CF approved annex5/Aprobadas y ajustadas por eco 19-01-2019/adjusted by eco/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/adjusted by eco/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Adjusted by Ecopetrol Jan 25 2019/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Adjusted by Eco Jan252019/'
path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/Approved by GCA adjusted by GCA(SAI)/'
# path_to_folder = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Cashflow Fix/test 2/'
# cashFlowFiles = []

# Get column numbers
colNum_AX = xlLib.col2num('AX')     # oil price ($/BBL)
colNum_AY = xlLib.col2num('AY')     # oil offset calidad price ($/BBL)
colNum_AZ = xlLib.col2num('AZ')     # oil offset transporte price ($/BBL)
colNum_AQ = xlLib.col2num('AQ')     # net oil volume (BBL)
colNum_AR = xlLib.col2num('AR')     # fuel oil volume (BBL)
colNum_AS = xlLib.col2num('AS')     # net gas volume (MCF)
colNum_AO = xlLib.col2num('AO')     # royalties gas volume (MCF)
colNum_BA = xlLib.col2num('BA')     # gas calorific factor (MMBTU/MCF)
colNum_BB = xlLib.col2num('BB')     # gas price ($/MMBTU)
colNum_AU = xlLib.col2num('AU')     # propane volume (BBL)
colNum_AV = xlLib.col2num('AV')     # butane volume (BBL)
colNum_AW = xlLib.col2num('AW')     # gasoline volume (BBL)
colNum_AP = xlLib.col2num('AP')     # royalties NGL volume (BBL)
colNum_BC = xlLib.col2num('BC')     # propane price ($/BBL)
colNum_BD = xlLib.col2num('BD')     # butane price ($/BBL)
colNum_BE = xlLib.col2num('BE')     # gasoline price ($/BBL)
colNum_BF = xlLib.col2num('BF')     # ECP revenue after royalties ($)

# arrays to store
cashFlowList = [file for file in os.listdir(path_to_folder) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_folder, file))]
numFields = len(cashFlowList)
oilPrices = np.zeros((2055-2018,5,numFields))
oilOff1 = np.zeros((2055-2018,5,numFields))
oilOff2 = np.zeros((2055-2018,5,numFields))
gasPrices = np.zeros((2055-2018,5,numFields))
propanePrices = np.zeros((2055-2018,5,numFields))
butanePrices = np.zeros((2055-2018,5,numFields))
gasolPrices = np.zeros((2055-2018,5,numFields))
calo_fac = np.zeros((2055-2018, 5,numFields))
netOil = np.zeros((2055-2018,5,numFields))
fuelOil = np.zeros((2055-2018,5,numFields))
netGas = np.zeros((2055-2018,5,numFields))
royaltiesGas = np.zeros((2055-2018,5,numFields))
netProp = np.zeros((2055-2018,5,numFields))
netBut = np.zeros((2055-2018,5,numFields))
netGasol = np.zeros((2055-2018,5,numFields))
royaltiesNGL = np.zeros((2055-2018,5,numFields))
netIncome = np.zeros((2055-2018,5,numFields))
calcIncome = np.zeros(netIncome.shape)
calcOilIncome = np.zeros(netIncome.shape)
calcGasIncome = np.zeros(netIncome.shape)
royaltiesGasDed = np.zeros(netIncome.shape)
calcNGLIncome = np.zeros(netIncome.shape)
royaltiesNGLDed = np.zeros(netIncome.shape)
isMatch_income = np.full((2055-2018,5,numFields), True)
avg_oilPrice = np.zeros((2055-2018,9))
avg_gasPrice = np.zeros((2055-2018,9))
avg_NGLPrice = np.zeros((2055-2018,9))

# # for loop to get all the data
#
# for i, file in enumerate(cashFlowList):
#     startTime = time.time()                                             # start timer for loop duration (DEBUG)
#     cashFlowFilePath = os.path.join(path_to_folder, file)                  # get cash flow workbook file path
#     wb = openpyxl.load_workbook(cashFlowFilePath, data_only=True)       # load workbook (read-only mode, faster)
#     ws_PDP = wb['Data CF-PDP']
#     ws_PNP = wb['Data CF-PNP']
#     ws_PND = wb['Data CF-PND']
#     ws_PRB = wb['Data CF-PRB']
#     ws_PS = wb['Data CF-PS']
#     wks = [ws_PDP, ws_PNP, ws_PND, ws_PRB, ws_PS]
#
#     # gather data from all the sheets in the cash flow
#     for k in range(len(wks)):
#         for j in range(36, 73):
#             netOil[j-36, k, i] = wks[k].cell(row=j, column=colNum_AQ).value
#             fuelOil[j-36, k, i] = wks[k].cell(row=j, column=colNum_AR).value
#             netGas[j-36, k, i] = wks[k].cell(row=j, column=colNum_AS).value
#             royaltiesGas[j-36, k, i] = wks[k].cell(row=j, column=colNum_AO).value
#             netProp[j-36, k, i] = wks[k].cell(row=j, column=colNum_AU).value
#             netBut[j-36, k, i] = wks[k].cell(row=j, column=colNum_AV).value
#             netGasol[j-36, k, i] = wks[k].cell(row=j, column=colNum_AW).value
#             royaltiesNGL[j-36, k, i] = wks[k].cell(row=j, column=colNum_AP).value
#             oilPrices[j-36, k, i] = wks[k].cell(row=j, column=colNum_AX).value
#             oilOff1[j-36, k, i] = wks[k].cell(row=j, column=colNum_AY).value
#             oilOff2[j-36, k, i] = wks[k].cell(row=j, column=colNum_AZ).value
#             gasPrices[j-36, k, i] = wks[k].cell(row=j, column=colNum_BB).value
#             calo_fac[j-36, k, i] = wks[k].cell(row=j, column=colNum_BA).value
#             propanePrices[j-36, k, i] = wks[k].cell(row=j, column=colNum_BC).value
#             butanePrices[j-36, k, i] = wks[k].cell(row=j, column=colNum_BD).value
#             gasolPrices[j-36, k, i] = wks[k].cell(row=j, column=colNum_BE).value
#             netIncome[j-36, k, i] = wks[k].cell(row=j, column=colNum_BF).value
#
#     # helpful message after saving workbook
#     print('Workbook ' + os.path.basename(cashFlowFilePath) + ' is extracted.')
#     stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
#     print('Time elapsed: ' + str(stopTime-startTime))
#     # ask for confirmation before proceeding to next file (DEBUG)
#     # wait = input('Press Enter to continue to the next workbook...')
#
# # save np arrays to disk to allow reloading them without having to run openpyxl again
# np.save('./NPy Test/netOil.npy', netOil)
# np.save('./NPy Test/fuelOil.npy', fuelOil)
# np.save('./NPy Test/netGas.npy', netGas)
# np.save('./NPy Test/royaltiesGas.npy', royaltiesGas)
# np.save('./NPy Test/netProp.npy', netProp)
# np.save('./NPy Test/netBut.npy', netBut)
# np.save('./NPy Test/netGasol.npy', netGasol)
# np.save('./NPy Test/royaltiesNGL.npy', royaltiesNGL)
# np.save('./NPy Test/oilPrices.npy', oilPrices)
# np.save('./NPy Test/oilOff1.npy', oilOff1)
# np.save('./NPy Test/oilOff2.npy', oilOff2)
# np.save('./NPy Test/gasPrices.npy', gasPrices)
# np.save('./NPy Test/calo_fac.npy', calo_fac)
# np.save('./NPy Test/propanePrices.npy', propanePrices)
# np.save('./NPy Test/butanePrices.npy', butanePrices)
# np.save('./NPy Test/gasolPrices.npy', gasolPrices)
# np.save('./NPy Test/netIncome.npy', netIncome)
#
# print('Numpy databases have been saved.')

################################################################################
# END OF FIRST PART OF CODE: LOADING DATA FROM EXCEL AND SAVING TO NUMPY DB
################################################################################

# load np data from stored databases
netOil = np.load('./NPy CF Data EcoAdj Jan252019/netOil.npy')
fuelOil = np.load('./NPy CF Data EcoAdj Jan252019/fuelOil.npy')
netGas = np.load('./NPy CF Data EcoAdj Jan252019/netGas.npy')
royaltiesGas = np.load('./NPy CF Data EcoAdj Jan252019/royaltiesGas.npy')
netProp = np.load('./NPy CF Data EcoAdj Jan252019/netProp.npy')
netBut = np.load('./NPy CF Data EcoAdj Jan252019/netBut.npy')
netGasol = np.load('./NPy CF Data EcoAdj Jan252019/netGasol.npy')
royaltiesNGL = np.load('./NPy CF Data EcoAdj Jan252019/royaltiesNGL.npy')
oilPrices = np.load('./NPy CF Data EcoAdj Jan252019/oilPrices.npy')
oilOff1 = np.load('./NPy CF Data EcoAdj Jan252019/oilOff1.npy')
oilOff2 = np.load('./NPy CF Data EcoAdj Jan252019/oilOff2.npy')
gasPrices = np.load('./NPy CF Data EcoAdj Jan252019/gasPrices.npy')
calo_fac = np.load('./NPy CF Data EcoAdj Jan252019/calo_fac.npy')
propanePrices = np.load('./NPy CF Data EcoAdj Jan252019/propanePrices.npy')
butanePrices = np.load('./NPy CF Data EcoAdj Jan252019/butanePrices.npy')
gasolPrices = np.load('./NPy CF Data EcoAdj Jan252019/gasolPrices.npy')
netIncome = np.load('./NPy CF Data EcoAdj Jan252019/netIncome.npy')

# Convert NaN to 0
netOil[np.isnan(netOil)] = 0
fuelOil[np.isnan(fuelOil)] = 0
netGas[np.isnan(netGas)] = 0
royaltiesGas[np.isnan(royaltiesGas)] = 0
netProp[np.isnan(netProp)] = 0
netBut[np.isnan(netBut)] = 0
netGasol[np.isnan(netGasol)] = 0
royaltiesNGL[np.isnan(royaltiesNGL)] = 0
oilPrices[np.isnan(oilPrices)] = 0
oilOff1[np.isnan(oilOff1)] = 0
oilOff2[np.isnan(oilOff2)] = 0
gasPrices[np.isnan(gasPrices)] = 0
calo_fac[np.isnan(calo_fac)] = 0
propanePrices[np.isnan(propanePrices)] = 0
butanePrices[np.isnan(butanePrices)] = 0
gasolPrices[np.isnan(gasolPrices)] = 0
netIncome[np.isnan(netIncome)] = 0

# for loop for income calculations (each year, category, field)
for i in range(numFields):
    for k in range(5):
        for j in range(2055-2018):
            calcOilIncome[j, k, i] = oilPrices[j, k, i] * (netOil[j, k, i] - fuelOil[j, k, i])
            calcGasIncome[j, k, i] = gasPrices[j, k, i] * calo_fac[j, k, i] * netGas[j, k, i]
            calcNGLIncome[j, k, i] = propanePrices[j, k, i] * netProp[j, k, i] + \
                                     butanePrices[j, k, i] * netBut[j, k, i] + \
                                     gasolPrices[j, k, i] * netGasol[j, k, i]
            royaltiesGasDed[j, k, i] = gasPrices[j, k, i] * calo_fac[j, k, i] * royaltiesGas[j, k, i]
            royaltiesNGLDed[j, k, i] = (propanePrices[j, k, i] + butanePrices[j, k, i] + gasolPrices[j, k, i]) / 3 * \
                                       royaltiesNGL[j, k, i]
            calcGasIncome[j, k, i] -= royaltiesGasDed[j, k, i]
            calcNGLIncome[j, k, i] -= royaltiesNGLDed[j, k, i]
            # calcIncome[j, k, i] = calcOilIncome[j, k, i] + calcGasIncome[j, k, i] + calcNGLIncome[j, k, i]
            # if np.abs(calcIncome[j, k, i]-netIncome[j, k, i])/netIncome[j, k, i] * 100 <= 0.5:
            #     isMatch_income[j, k, i] = True
            # else:
            #     isMatch_income[j, k, i] = False

# calculate total for the 5 reserves categories (PDP, PNP, PND, PRB, PS)
totalOil = np.zeros((2055-2018, 9))
totalOilIncome = np.zeros(totalOil.shape)
totalGas = np.zeros(totalOil.shape)
totalGasIncome = np.zeros(totalOil.shape)
totalNGL = np.zeros(totalOil.shape)
totalNGLIncome = np.zeros(totalOil.shape)
totalIncome = np.zeros(totalOil.shape)
totalIncome_calc = np.zeros(totalOil.shape)
for l in range(5):
    for j in range(2055-2018):
        for i in range(numFields):
            totalOil[j, l] += (netOil[j, l, i] - fuelOil[j, l, i])
            totalOilIncome[j, l] += calcOilIncome[j, l, i]
            totalGas[j, l] += (netGas[j, l, i] - royaltiesGas[j, l, i])
            totalGasIncome[j, l] += calcGasIncome[j, l, i]
            totalNGL[j, l] += (netProp[j, l, i] + netBut[j, l, i] + netGasol[j, l, i] - royaltiesNGL[j, l, i])
            totalNGLIncome[j, l] += calcNGLIncome[j, l, i]
            totalIncome[j, l] += netIncome[j, l, i]
# for the PD, 1p, 2p, 3p categories
for j in range(2055-2018):
    totalOil[j, 5] = totalOil[j, 0] + totalOil[j, 1]
    totalOil[j, 6] = totalOil[j, 5] + totalOil[j, 2]
    totalOil[j, 7] = totalOil[j, 6] + totalOil[j, 3]
    totalOil[j, 8] = totalOil[j, 7] + totalOil[j, 4]
    totalOilIncome[j, 5] = totalOilIncome[j, 0] + totalOilIncome[j, 1]
    totalOilIncome[j, 6] = totalOilIncome[j, 5] + totalOilIncome[j, 2]
    totalOilIncome[j, 7] = totalOilIncome[j, 6] + totalOilIncome[j, 3]
    totalOilIncome[j, 8] = totalOilIncome[j, 7] + totalOilIncome[j, 4]
    totalGas[j, 5] = totalGas[j, 0] + totalGas[j, 1]
    totalGas[j, 6] = totalGas[j, 5] + totalGas[j, 2]
    totalGas[j, 7] = totalGas[j, 6] + totalGas[j, 3]
    totalGas[j, 8] = totalGas[j, 7] + totalGas[j, 4]
    totalGasIncome[j, 5] = totalGasIncome[j, 0] + totalGasIncome[j, 1]
    totalGasIncome[j, 6] = totalGasIncome[j, 5] + totalGasIncome[j, 2]
    totalGasIncome[j, 7] = totalGasIncome[j, 6] + totalGasIncome[j, 3]
    totalGasIncome[j, 8] = totalGasIncome[j, 7] + totalGasIncome[j, 4]
    totalNGL[j, 5] = totalNGL[j, 0] + totalNGL[j, 1]
    totalNGL[j, 6] = totalNGL[j, 5] + totalNGL[j, 2]
    totalNGL[j, 7] = totalNGL[j, 6] + totalNGL[j, 3]
    totalNGL[j, 8] = totalNGL[j, 7] + totalNGL[j, 4]
    totalNGLIncome[j, 5] = totalNGLIncome[j, 0] + totalNGLIncome[j, 1]
    totalNGLIncome[j, 6] = totalNGLIncome[j, 5] + totalNGLIncome[j, 2]
    totalNGLIncome[j, 7] = totalNGLIncome[j, 6] + totalNGLIncome[j, 3]
    totalNGLIncome[j, 8] = totalNGLIncome[j, 7] + totalNGLIncome[j, 4]
    totalIncome[j, 5] = totalIncome[j, 0] + totalIncome[j, 1]
    totalIncome[j, 6] = totalIncome[j, 5] + totalIncome[j, 2]
    totalIncome[j, 7] = totalIncome[j, 6] + totalIncome[j, 3]
    totalIncome[j, 8] = totalIncome[j, 7] + totalIncome[j, 4]
# total income calc and avg prices cals
err_income = np.zeros(totalIncome.shape)
for l in range(9):
    for j in range(2055-2018):
        totalIncome_calc[j, l] = totalOilIncome[j, l] + totalGasIncome[j, l] + totalNGLIncome[j, l]
        if totalIncome[j, l] != 0:
            err_income[j, l] = (totalIncome[j, l] - totalIncome_calc[j, l]) / totalIncome[j, l]
        else:
            err_income[j, l] = 0.0
        if totalOil[j, l] != 0:
            avg_oilPrice[j, l] = totalOilIncome[j, l] * 1.0 / totalOil[j, l]
        else:
            avg_oilPrice[j, l] = 0.0
        if totalGas[j, l] != 0:
            avg_gasPrice[j, l] = totalGasIncome[j, l] * 1.0 / totalGas[j, l]
        else:
            avg_gasPrice[j, l] = 0.0
        if totalNGL[j, l] != 0:
            avg_NGLPrice[j, l] = totalNGLIncome[j, l] * 1.0 / totalNGL[j, l]
        else:
            avg_NGLPrice[j, l] = 0.0

# Write output to excel sheet
wb_out = openpyxl.Workbook(write_only=True)
ws0 = wb_out.create_sheet('PDP')
ws1 = wb_out.create_sheet('PNP')
ws2 = wb_out.create_sheet('PND')
ws3 = wb_out.create_sheet('PRB')
ws4 = wb_out.create_sheet('PS')
ws5 = wb_out.create_sheet('PD')
ws6 = wb_out.create_sheet('1P')
ws7 = wb_out.create_sheet('2P')
ws8 = wb_out.create_sheet('3P')
wks_out = [ws0, ws1, ws2, ws3, ws4, ws5, ws6, ws7, ws8]
years = np.arange(2019,2056)

for l in range(9):
    wks_out[l].append(['', 'Oil Price', 'Gas Price', 'NGL Price', 'Calc. Revenue', 'ECP Revenue', 'Error in Revenue'])
    wks_out[l].append(['Year', 'U.S.$/bbl', 'U.S.$/Mcf', 'U.S.$/bbl', 'MM U.S.$', 'MM U.S.$', ''])
    for j in range(2055-2018):
        wks_out[l].append([years[j], avg_oilPrice[j, l], avg_gasPrice[j,l], \
                           avg_NGLPrice[j, l], totalIncome_calc[j,l]/1e6, \
                           totalIncome[j, l]/1e6, err_income[j, l]])

wb_out.save('Average_Prices.xlsx')
print('Successfully done!')
