################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2018
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
import numpy as np

path_to_reserves = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Reserves Tables/Cashflows/'

# # Allocate arrays to store
# oil = np.zeros((64, 4, 5))
# fuel_oil = np.zeros(oil.shape)
# ngl = np.zeros(oil.shape)
# gas = np.zeros(oil.shape)
# fuel_gas = np.zeros(oil.shape)
# fieldNames = []
# cashFlowList = [file for file in os.listdir(path_to_reserves) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_reserves, file))]
# fieldNames = [cashFlow[:cashFlow.find('_')].upper() for cashFlow in cashFlowList]
# fileSortOrder = np.argsort(fieldNames)
# cashFlowList = [cashFlowList[i] for i in fileSortOrder]
#
# # for loop to get all the reserves data from cash flows
# for i, file in enumerate(cashFlowList):
#     startTime = time.time()                                             # start timer for loop duration (DEBUG)
#     cashFlowFilePath = os.path.join(path_to_reserves, file)                  # get cash flow workbook file path
#     wb = openpyxl.load_workbook(cashFlowFilePath, data_only=True)       # load workbook (read-only mode, faster)
#     ws_PDP = wb['Data CF PDP']
#     ws_PNP = wb['Data CF PNP']
#     ws_PND = wb['Data CF PND']
#     ws_PRB = wb['Data CF PRB']
#     ws_PS = wb['Data CF PS']
#     wks = [ws_PDP, ws_PNP, ws_PND, ws_PRB, ws_PS]
#
#     # get field name
#     fieldNames.append(wks[0].cell(row=3, column=2).value)
#
#     # gather data from all the sheets in the cash flow
#     for k in range(len(wks)):   # categories of reserves (producing, non-producing, undeveloped, probable, possible)
#         # Oil
#         oil[i, 0, k] = wks[k].cell(row=74, column=xlLib.col2num('L')).value / 1000     # gross tech vol, Oil (Mbbl)
#         oil[i, 1, k] = wks[k].cell(row=74, column=xlLib.col2num('S')).value / 1000     # gross field res, Oil (Mbbl)
#         oil[i, 2, k] = wks[k].cell(row=74, column=xlLib.col2num('AG')).value / 1000    # gross res to ECP, Oil (Mbbl)
#         oil[i, 3, k] = wks[k].cell(row=74, column=xlLib.col2num('AQ')).value / 1000    # net res to ECP, Oil (Mbbl)
#         # Fuel Oil
#         fuel_oil[i, 0, k] = wks[k].cell(row=74, column=xlLib.col2num('M')).value / 1000     # gross tech vol, Fuel Oil (Mbbl)
#         fuel_oil[i, 1, k] = wks[k].cell(row=74, column=xlLib.col2num('T')).value / 1000     # gross field res, Fuel Oil (Mbbl)
#         fuel_oil[i, 2, k] = wks[k].cell(row=74, column=xlLib.col2num('AH')).value / 1000    # gross res to ECP, Fuel Oil (Mbbl)
#         fuel_oil[i, 3, k] = wks[k].cell(row=74, column=xlLib.col2num('AR')).value / 1000    # net res to ECP, Fuel Oil (Mbbl)
#         # NGLs
#         ngl[i, 0, k] = (wks[k].cell(row=74, column=xlLib.col2num('P')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('Q')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('R')).value) / 1000     # gross tech vol, NGLs (Mbbl)
#         ngl[i, 1, k] = (wks[k].cell(row=74, column=xlLib.col2num('W')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('X')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('Y')).value) / 1000     # gross field res, NGLs (Mbbl)
#         ngl[i, 2, k] = (wks[k].cell(row=74, column=xlLib.col2num('AK')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('AL')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('AM')).value) / 1000    # gross res to ECP, NGLs (Mbbl)
#         ngl[i, 3, k] = (wks[k].cell(row=74, column=xlLib.col2num('AU')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('AV')).value + \
#                         wks[k].cell(row=74, column=xlLib.col2num('AW')).value) / 1000    # net res to ECP, NGLs (Mbbl)
#         # Gas
#         gas[i, 0, k] = wks[k].cell(row=74, column=xlLib.col2num('N')).value / 1000     # gross tech vol, Gas (MMcf)
#         gas[i, 1, k] = wks[k].cell(row=74, column=xlLib.col2num('U')).value / 1000     # gross field res, Gas (MMcf)
#         gas[i, 2, k] = wks[k].cell(row=74, column=xlLib.col2num('AI')).value / 1000    # gross res to ECP, Gas (MMcf)
#         gas[i, 3, k] = wks[k].cell(row=74, column=xlLib.col2num('AS')).value / 1000    # net res to ECP, Gas (MMcf)
#         # Fuel Gas
#         fuel_gas[i, 0, k] = wks[k].cell(row=74, column=xlLib.col2num('O')).value / 1000     # gross tech vol, Fuel Gas (MMcf)
#         fuel_gas[i, 1, k] = wks[k].cell(row=74, column=xlLib.col2num('V')).value / 1000     # gross field res, Fuel Gas (MMcf)
#         fuel_gas[i, 2, k] = wks[k].cell(row=74, column=xlLib.col2num('AJ')).value / 1000    # gross res to ECP, Fuel Gas (MMcf)
#         fuel_gas[i, 3, k] = wks[k].cell(row=74, column=xlLib.col2num('AT')).value / 1000    # net res to ECP, Fuel Gas (MMcf)
#
#     # helpful message after saving workbook
#     print('Workbook ' + os.path.basename(cashFlowFilePath) + ' from cash flows is extracted.')
#     stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
#     print('Time elapsed: ' + str(stopTime-startTime))
#
# # Write extracted data to Numpy databases for use below
# np.save('./reserves/oil.npy', oil)
# np.save('./reserves/fuel_oil.npy', fuel_oil)
# np.save('./reserves/ngl.npy', ngl)
# np.save('./reserves/gas.npy', gas)
# np.save('./reserves/fuel_gas.npy', fuel_gas)
# np.save('./reserves/fieldNames.npy', np.array(fieldNames))

# Reload reserves data from Numpy databases
oil = np.load('./reserves/oil.npy')
fuel_oil = np.load('./reserves/fuel_oil.npy')
ngl = np.load('./reserves/ngl.npy')
gas = np.load('./reserves/gas.npy')
fuel_gas = np.load('./reserves/fuel_gas.npy')
fieldNames = np.load('./reserves/fieldNames.npy')
fieldNames = fieldNames[64:128]     # fix to deal with a mistake in extracting field names in previous part of code

# Write all data into new workbook for use in creating reserves tables
wb = openpyxl.Workbook()
out_filename = 'Reserves_Tables.xlsx'

ws_pdp = wb.create_sheet(title='PDP')
ws_pnp = wb.create_sheet(title='PNP')
ws_pnd = wb.create_sheet(title='PND')
ws_prb = wb.create_sheet(title='PRB')
ws_ps = wb.create_sheet(title='PS')
wks = [ws_pdp, ws_pnp, ws_pnd, ws_prb, ws_ps]

for k in range(len(wks)):   # categories of reserves (producing, non-producing, undeveloped, probable, possible)
    ws = wks[k]
    # Create headers in each sheet
    ws['A3'] = 'Field'

    ws['B1'] = 'Gross Technical Volumes'
    ws['B2'] = 'Oil'
    ws['B3'] = 'MBbl'
    ws['C2'] = 'Fuel Oil'
    ws['C3'] = 'MBbl'
    ws['D2'] = 'NGLs'
    ws['D3'] = 'MBbl'
    ws['E2'] = 'Gas'
    ws['E3'] = 'MMscf'
    ws['F2'] = 'Fuel Gas'
    ws['F3'] = 'MMscf'

    ws['H1'] = 'Gross Field Volumes'
    ws['H2'] = 'Oil'
    ws['H3'] = 'MBbl'
    ws['I2'] = 'Fuel Oil'
    ws['I3'] = 'MBbl'
    ws['J2'] = 'NGLs'
    ws['J3'] = 'MBbl'
    ws['K2'] = 'Gas'
    ws['K3'] = 'MMscf'
    ws['L2'] = 'Fuel Gas'
    ws['L3'] = 'MMscf'

    ws['N1'] = 'Gross Field Volumes (Ecopetrol Interest)'
    ws['N2'] = 'Oil'
    ws['N3'] = 'MBbl'
    ws['O2'] = 'Fuel Oil'
    ws['O3'] = 'MBbl'
    ws['P2'] = 'NGLs'
    ws['P3'] = 'MBbl'
    ws['Q2'] = 'Gas'
    ws['Q3'] = 'MMscf'
    ws['R2'] = 'Fuel Gas'
    ws['R3'] = 'MMscf'

    ws['T1'] = 'Reserves Net to Ecopetrol Interest'
    ws['T2'] = 'Oil'
    ws['T3'] = 'MBbl'
    ws['U2'] = 'Fuel Oil'
    ws['U3'] = 'MBbl'
    ws['V2'] = 'NGLs'
    ws['V3'] = 'MBbl'
    ws['W2'] = 'Gas'
    ws['W3'] = 'MMscf'
    ws['X2'] = 'Fuel Gas'
    ws['X3'] = 'MMscf'

    for i in range(fieldNames.shape[0]):    # 64 fields
        # field name
        _ = ws.cell(column=xlLib.col2num('A'), row=i+4, value=fieldNames[i])     # Field
        # oil
        _ = ws.cell(column=xlLib.col2num('B'), row=i+4, value=oil[i, 0, k])     # GTV, oil
        _ = ws.cell(column=xlLib.col2num('H'), row=i+4, value=oil[i, 1, k])     # GFV, oil
        _ = ws.cell(column=xlLib.col2num('N'), row=i+4, value=oil[i, 2, k])     # GFV to ECP, oil
        _ = ws.cell(column=xlLib.col2num('T'), row=i+4, value=oil[i, 3, k])     # Net to ECP, oil
        # fuel oil
        _ = ws.cell(column=xlLib.col2num('C'), row=i+4, value=fuel_oil[i, 0, k])     # GTV, fuel oil
        _ = ws.cell(column=xlLib.col2num('I'), row=i+4, value=fuel_oil[i, 1, k])     # GFV, fuel oil
        _ = ws.cell(column=xlLib.col2num('O'), row=i+4, value=fuel_oil[i, 2, k])     # GFV to ECP, fuel oil
        _ = ws.cell(column=xlLib.col2num('U'), row=i+4, value=fuel_oil[i, 3, k])     # Net to ECP, fuel oil
        # NGLs
        _ = ws.cell(column=xlLib.col2num('D'), row=i+4, value=ngl[i, 0, k])     # GTV, ngl
        _ = ws.cell(column=xlLib.col2num('J'), row=i+4, value=ngl[i, 1, k])     # GFV, ngl
        _ = ws.cell(column=xlLib.col2num('P'), row=i+4, value=ngl[i, 2, k])     # GFV to ECP, ngl
        _ = ws.cell(column=xlLib.col2num('V'), row=i+4, value=ngl[i, 3, k])     # Net to ECP, ngl
        # gas
        _ = ws.cell(column=xlLib.col2num('E'), row=i+4, value=gas[i, 0, k])     # GTV, gas
        _ = ws.cell(column=xlLib.col2num('K'), row=i+4, value=gas[i, 1, k])     # GFV, gas
        _ = ws.cell(column=xlLib.col2num('Q'), row=i+4, value=gas[i, 2, k])     # GFV to ECP, gas
        _ = ws.cell(column=xlLib.col2num('W'), row=i+4, value=gas[i, 3, k])     # Net to ECP, gas
        # fuel gas
        _ = ws.cell(column=xlLib.col2num('F'), row=i+4, value=fuel_gas[i, 0, k])     # GTV, fuel gas
        _ = ws.cell(column=xlLib.col2num('L'), row=i+4, value=fuel_gas[i, 1, k])     # GFV, fuel gas
        _ = ws.cell(column=xlLib.col2num('R'), row=i+4, value=fuel_gas[i, 2, k])     # GFV to ECP, fuel gas
        _ = ws.cell(column=xlLib.col2num('X'), row=i+4, value=fuel_gas[i, 3, k])     # Net to ECP, fuel gas

# save workbook
wb.save(filename = out_filename)
