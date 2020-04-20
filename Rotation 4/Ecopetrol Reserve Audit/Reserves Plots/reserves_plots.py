################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2018
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import datetime
import matplotlib.dates as mdates
import matplotlib.backends.backend_pdf

path_to_history = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Reserves Plots/'
path_to_reserves = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Reserves Plots/Cashflows/'

# Get historical data into arrays
years_hist = np.arange(1936, 2019)
years_hist_dt = []
for i in range(len(years_hist)):
    years_hist_dt.append(datetime.datetime(years_hist[i], 12, 31))
yearsFmt = mdates.DateFormatter('%Y')
years_loc = mdates.YearLocator()   # every year
months_loc = mdates.MonthLocator()  # every month
# oil_hist = []       # oil production (bopd)
# gas_hist = []       # gas production (mscfd)
# water_hist = []     # water production (bwpd)
# histFilePath = os.path.join(path_to_history, 'historical_production.xlsx')
# wb = openpyxl.load_workbook(histFilePath, data_only=True)
# ws = wb['Sheet1']
# row = 3     # starting row
# for i in range(len(years_hist)):
#     oil_hist_row = []
#     gas_hist_row = []
#     water_hist_row = []
#     col = 2     # starting column
#     for j in range(64):
#         oil_hist_row.append(ws.cell(row=row, column=col).value)
#         col += 1
#         gas_hist_row.append(ws.cell(row=row, column=col).value)
#         col += 1
#         water_hist_row.append(ws.cell(row=row, column=col).value)
#         col += 1
#     oil_hist.append(oil_hist_row)
#     gas_hist.append(gas_hist_row)
#     water_hist.append(water_hist_row)
#     row += 1
#
# # convert to numpy arrays
# oil_hist = np.array(oil_hist, dtype=float)
# gas_hist = np.array(gas_hist, dtype=float)
# water_hist = np.array(water_hist, dtype=float)
#
# # get field names
# fields = []
# col = 2
# for i in range(64):
#     fields.append(ws.cell(row=1, column=col).value)
#     col += 3
# fields = np.array(fields)
#
# # save historical data to numpy db for use later
# np.save('./history/oil_hist.npy', oil_hist)
# np.save('./history/gas_hist.npy', gas_hist)
# np.save('./history/water_hist.npy', water_hist)
# np.save('./fields.npy', fields)
#
# print('Historical data has been extracted.')


# Get reserves data into arrays
years = np.arange(2019, 2056)
years_dt = []
for i in range(len(years)):
    years_dt.append(datetime.datetime(years[i], 1, 1))
# oil = np.zeros((2056-2019, 64, 9))
# fuel_oil = np.zeros(oil.shape)
# gas = np.zeros(oil.shape)
# fuel_gas = np.zeros(oil.shape)
# ngl1 = np.zeros(oil.shape)
# ngl2 = np.zeros(oil.shape)
# ngl3 = np.zeros(oil.shape)
# cashFlowList = [file for file in os.listdir(path_to_reserves) if file.endswith('.xlsx') and os.path.isfile(os.path.join(path_to_reserves, file))]
# fieldNames = [cashFlow[:cashFlow.find('_')].upper() for cashFlow in cashFlowList]
# fileSortOrder = np.argsort(fieldNames)
# cashFlowList = [cashFlowList[i] for i in fileSortOrder]
#
# # for loop to get all the reserves data (gross production volumes) from cash flows
# for i, file in enumerate(cashFlowList):
#     startTime = time.time()                                             # start timer for loop duration (DEBUG)
#     cashFlowFilePath = os.path.join(path_to_reserves, file)                  # get cash flow workbook file path
#     wb = openpyxl.load_workbook(cashFlowFilePath, data_only=True)       # load workbook (read-only mode, faster)
#     ws_PDP = wb['Data CF PDP']
#     ws_PNP = wb['Data CF PNP']
#     ws_PND = wb['Data CF PND']
#     ws_PRB = wb['Data CF PRB']
#     ws_PS = wb['Data CF PS']
#     wks = [ws_PDP, ws_PNP, ws_PND, ws_PRB, ws_PS]
#
#     # gather data from all the sheets in the cash flow
#     for k in range(len(wks)):
#         row = 36    # starting row (year 2019)
#         for j in range(len(years)):
#             oil[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('S')).value
#             fuel_oil[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('T')).value
#             gas[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('U')).value
#             fuel_gas[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('V')).value
#             ngl1[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('W')).value
#             ngl2[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('X')).value
#             ngl3[j, i, k] = wks[k].cell(row=row, column=xlLib.col2num('Y')).value
#             row += 1
#
#     # helpful message after saving workbook
#     print('Workbook ' + os.path.basename(cashFlowFilePath) + ' from cash flows is extracted.')
#     stopTime = time.time()                                              # stop timer for loop duration (DEBUG)
#     print('Time elapsed: ' + str(stopTime-startTime))
#     # ask for confirmation before proceeding to next file (DEBUG)
#     # wait = input('Press Enter to continue to the next workbook...')
#
# # convert to float (None to nan?)
# oil = np.array(oil, dtype=float)
# oil[np.isnan(oil)] = 0.0
# fuel_oil = np.array(fuel_oil, dtype=float)
# fuel_oil[np.isnan(fuel_oil)] = 0.0
# oil = oil + fuel_oil        # include fuel oil in oil reserves reporting
# gas = np.array(gas, dtype=float)
# gas[np.isnan(gas)] = 0.0
# fuel_gas = np.array(fuel_gas, dtype=float)
# fuel_gas[np.isnan(fuel_gas)] = 0.0
# gas = gas + fuel_gas        # include fuel gas in gas reserves reporting
# ngl1 = np.array(ngl1, dtype=float)
# ngl1[np.isnan(ngl1)] = 0.0
# ngl2 = np.array(ngl2, dtype=float)
# ngl2[np.isnan(ngl2)] = 0.0
# ngl3 = np.array(ngl3, dtype=float)
# ngl3[np.isnan(ngl3)] = 0.0
# ngl = ngl1 + ngl2 + ngl3
#
# # save np arrays to disk to allow reloading them without having to run openpyxl again
# np.save('./reserves/oil.npy', oil)
# np.save('./reserves/fuel_oil.npy', fuel_oil)
# np.save('./reserves/gas.npy', gas)
# np.save('./reserves/fuel_gas.npy', fuel_gas)
# np.save('./reserves/ngl.npy', ngl)
#
# print('Numpy databases for reserves data from cash flows have been saved.')


# Reload historical and field name data (already in rates)
oil_hist = np.load('./history/oil_hist.npy')
gas_hist = np.load('./history/gas_hist.npy')
water_hist = np.load('./history/water_hist.npy')
fields = np.load('./fields.npy')

# Reload reserves data (volumes) and convert to rates
oil = np.load('./reserves/oil.npy') / 365
gas = np.load('./reserves/gas.npy') / 365
ngl = np.load('./reserves/ngl.npy') / 365

# calculations for PD, 1P, 2P and 3P categories
for i in range(64):
    # PD
    oil[:, i, 5] = oil[:, i, 0] + oil[:, i, 1]
    gas[:, i, 5] = gas[:, i, 0] + gas[:, i, 1]
    ngl[:, i, 5] = ngl[:, i, 0] + ngl[:, i, 1]
    # 1P
    oil[:, i, 6] = oil[:, i, 5] + oil[:, i, 2]
    gas[:, i, 6] = gas[:, i, 5] + gas[:, i, 2]
    ngl[:, i, 6] = ngl[:, i, 5] + ngl[:, i, 2]
    # 2P
    oil[:, i, 7] = oil[:, i, 6] + oil[:, i, 3]
    gas[:, i, 7] = gas[:, i, 6] + gas[:, i, 3]
    ngl[:, i, 7] = ngl[:, i, 6] + ngl[:, i, 3]
    # 3P
    oil[:, i, 8] = oil[:, i, 7] + oil[:, i, 4]
    gas[:, i, 8] = gas[:, i, 7] + gas[:, i, 4]
    ngl[:, i, 8] = ngl[:, i, 7] + ngl[:, i, 4]

#------------------------------------------------------------------------------
# PLOTTING
#-------------------------------------------------------------------------------
arial_font = fm.FontProperties(family='arial', size=8)
# arial_font.set_family('arial')

min_oil = np.nanmin(np.array([np.nanmin(oil_hist), np.nanmin(oil)]))
if min_oil <= 0:
    min_oil = 1
max_oil = np.nanmax(np.array([np.nanmax(oil_hist), np.nanmax(oil)])) * 1.05
min_gas = np.nanmin(np.array([np.nanmin(gas_hist), np.nanmin(gas)]))
if min_gas <= 0:
    min_gas = 1
max_gas = np.nanmax(np.array([np.nanmax(gas_hist), np.nanmax(gas)])) * 1.05
min_ngl = np.nanmin(ngl)
if min_ngl <= 0:
    min_ngl = 1
max_ngl = np.nanmax(ngl) * 1.05

for i in range(len(fields)):
    # output PDF file name and setting
    out_pdf = './reserves charts/{}.pdf'.format(fields[i])
    pdf = matplotlib.backends.backend_pdf.PdfPages(out_pdf)

    # initial modification for only plotting independent profiles
    # e.g. if 1P = 2P, then only plot 1P but show legend for 2P as well

    # OIL
    # case 1: 3p = 2p
    if (oil[:, i, 8] == oil[:, i, 7]).all():
        oil[:, i, 8] = np.nan
    # case 2: 2p = 1p
    if (oil[:, i, 7] == oil[:, i, 6]).all():
        oil[:, i, 7] = np.nan
    # case 3: 1p = PDP
    if (oil[:, i, 6] == oil[:, i, 0]).all():
        oil[:, i, 6] = np.nan

    # GAS
    # case 1: 3p = 2p
    if (gas[:, i, 8] == gas[:, i, 7]).all():
        gas[:, i, 8] = np.nan
    # case 2: 2p = 1p
    if (gas[:, i, 7] == gas[:, i, 6]).all():
        gas[:, i, 7] = np.nan
    # case 3: 1p = PDP
    if (gas[:, i, 6] == gas[:, i, 0]).all():
        gas[:, i, 6] = np.nan

    # NGL
    # case 1: 3p = 2p
    if (ngl[:, i, 8] == ngl[:, i, 7]).all():
        ngl[:, i, 8] = np.nan
    # case 2: 2p = 1p
    if (ngl[:, i, 7] == ngl[:, i, 6]).all():
        ngl[:, i, 7] = np.nan
    # case 3: 1p = PDP
    if (ngl[:, i, 6] == ngl[:, i, 0]).all():
        ngl[:, i, 6] = np.nan

    # oil plots
    fig, ax = plt.subplots()
    ax.semilogy(years_hist_dt, oil_hist[:,i], 'k--', label='Historical')
    ax.semilogy(years_dt, oil[:, i, 0], '-', label='PDP')
    ax.semilogy(years_dt, oil[:, i, 6], '-', label='1P')
    ax.semilogy(years_dt, oil[:, i, 7], '-', label='2P')
    ax.semilogy(years_dt, oil[:, i, 8], '-', label='3P')
    ax.set_ylim(min_oil, max_oil)
    ax.yaxis.grid(which="major", linestyle='-', linewidth=1)
    ax.yaxis.grid(which="minor", linestyle='-', linewidth=0.5)
    # ax.xaxis.set_major_locator(years_loc)
    # ax.xaxis.set_major_formatter(yearsFmt)
    # ax.xaxis.set_minor_locator(years_loc)
    ax.set_xlabel('Year', fontname='Arial', fontsize=8)
    ax.set_ylabel('Oil Rate (bopd)', fontname='Arial', fontsize=8)
    ax.set_title('{0} Field, Colombia\nGross (100% WI) Oil Volumes as of December 31, 2018\n(oil rate includes fuel oil)'.format(fields[i]), fontname='Arial', fontsize=8)
    for tick in ax.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax.get_yticklabels():
        tick.set_fontname("Arial")
    ax.legend(loc='best', prop=arial_font)

    # Date labels look better
    fig.autofmt_xdate()

    # plt.show()
    if np.nansum(oil[:, i, :]) + np.nansum(oil_hist[:, i]) > 0 and \
       ((oil[:, i, :] > 1).any() or (oil_hist[:, i] > 1).any()):
        # plt.savefig('./plots/{}_PLOT1_oil.png'.format(fields[i]), dpi=300)
        pdf.savefig(fig, dpi=300)
    plt.close()

    # gas plots
    fig, ax = plt.subplots()
    ax.semilogy(years_hist_dt, gas_hist[:,i], 'k--', label='Historical')
    ax.semilogy(years_dt, gas[:, i, 0], '-', label='PDP')
    ax.semilogy(years_dt, gas[:, i, 6], '-', label='1P')
    ax.semilogy(years_dt, gas[:, i, 7], '-', label='2P')
    ax.semilogy(years_dt, gas[:, i, 8], '-', label='3P')
    ax.set_ylim(min_gas, max_gas)
    ax.yaxis.grid(which="major", linestyle='-', linewidth=1)
    ax.yaxis.grid(which="minor", linestyle='-', linewidth=0.5)
    # ax.xaxis.set_major_locator(years_loc)
    # ax.xaxis.set_major_formatter(yearsFmt)
    # ax.xaxis.set_minor_locator(years_loc)
    ax.set_xlabel('Year', fontname='Arial', fontsize=8)
    ax.set_ylabel('Gas Rate (Mscfd)', fontname='Arial', fontsize=8)
    ax.set_title('{} Field, Colombia\nGross (100% WI) Gas Volumes as of December 31, 2018\n(gas rate includes fuel gas)'.format(fields[i]), fontname='Arial', fontsize=8)
    for tick in ax.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax.get_yticklabels():
        tick.set_fontname("Arial")
    ax.legend(loc='best', prop=arial_font)

    # Date labels look better
    fig.autofmt_xdate()

    # plt.show()
    if np.nansum(gas[:, i, :]) + np.nansum(gas_hist[:, i]) > 0 and \
       ((gas[:, i, :] > 1).any() or (gas_hist[:, i] > 1).any()):
        # plt.savefig('./plots/{}_PLOT2_gas.png'.format(fields[i]), dpi=300)
        pdf.savefig(fig, dpi=300)
    plt.close()

    # ngl plots
    fig, ax = plt.subplots()
    ax.semilogy(years_dt, ngl[:, i, 0], '-', label='PDP')
    ax.semilogy(years_dt, ngl[:, i, 6], '-', label='1P')
    ax.semilogy(years_dt, ngl[:, i, 7], '-', label='2P')
    ax.semilogy(years_dt, ngl[:, i, 8], '-', label='3P')
    ax.set_ylim(min_ngl, max_ngl)
    ax.yaxis.grid(which="major", linestyle='-', linewidth=1)
    ax.yaxis.grid(which="minor", linestyle='-', linewidth=0.5)
    # ax.xaxis.set_major_locator(years_loc)
    # ax.xaxis.set_major_formatter(yearsFmt)
    # ax.xaxis.set_minor_locator(years_loc)
    ax.set_xlabel('Year', fontname='Arial', fontsize=8)
    ax.set_ylabel('Liquids Rate (bpd)', fontname='Arial', fontsize=8)
    ax.set_title('{} Field, Colombia\nGross (100% WI) NGL Volumes as of December 31, 2018'.format(fields[i]), fontname='Arial', fontsize=8)
    for tick in ax.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax.get_yticklabels():
        tick.set_fontname("Arial")
    ax.legend(loc='best', prop=arial_font)

    # Date labels look better
    fig.autofmt_xdate()

    # plt.show()
    if np.nansum(ngl[:, i, :]) > 0 and (ngl[:, i, :] > 1).any():
        # plt.savefig('./plots/{}_PLOT3_NGL.png'.format(fields[i]), dpi=300)
        pdf.savefig(fig, dpi=300)
    plt.close()

    # close pdf for field
    pdf.close()
