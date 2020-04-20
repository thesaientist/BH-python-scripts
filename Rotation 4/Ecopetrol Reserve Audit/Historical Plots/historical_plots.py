################################################################################
##
##    Property of Baker Hughes, a GE Company
##    Copyright 2019
##    Author: Sai Uppati
##    saipranav.uppati@bhge.com
##
################################################################################

import os
import openpyxl
import openpyxl_lib as xlLib
import time
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
import datetime
import matplotlib.dates as mdates
import matplotlib.backends.backend_pdf

def CreateHistoryPlot(field=None, isInj=False):
    # Field and path
    if field is None:
        FIELD_NAME = 'CARACARA'
    else:
        FIELD_NAME = field
    path_to_history = 'C:/Users/212566876/Box Sync/Rotation 4/Ecopetrol Reserve Audit/Historical Plots'

    # Get historical data into arrays
    # years_hist = np.arange(1936, 2019)
    # years_hist_dt = []
    yearsFmt = mdates.DateFormatter('%Y')
    years_loc = mdates.YearLocator()   # every year
    months_loc = mdates.MonthLocator()  # every month
    months_hist = []    # month/year of production
    # production
    oil_hist = []       # oil production (bopd)
    gas_hist = []       # gas production (mscfd)
    water_hist = []     # water production (bwpd)
    histFilePath = os.path.join(path_to_history, '{}.xlsx'.format(FIELD_NAME))
    wb = openpyxl.load_workbook(histFilePath, data_only=True)
    ws_prod = wb['Prod']
    row = 2     # starting row
    isNextMonth = True
    while isNextMonth:
        oil_hist.append(ws_prod.cell(row=row, column=6).value)
        gas_hist.append(ws_prod.cell(row=row, column=8).value)
        water_hist.append(ws_prod.cell(row=row, column=7).value)
        month_val = str(ws_prod.cell(row=row, column=1).value)
        months_hist.append(datetime.datetime.strptime(month_val[:10], '%Y-%m-%d'))
        row += 1
        if ws_prod.cell(row=row, column=1).value is None:
            isNextMonth = False
    # injection
    if isInj:
        monthsinj_hist = []    # month/year of injection
        waterinj_hist = []     # water injection (bwpd)
        ws_inj = wb['Inj']
        row = 2     # starting row
        isNextMonth = True
        while isNextMonth:
            waterinj_hist.append(ws_inj.cell(row=row, column=4).value)
            month_val = str(ws_inj.cell(row=row, column=1).value)
            monthsinj_hist.append(datetime.datetime.strptime(month_val[:10], '%Y-%m-%d'))
            row += 1
            if ws_inj.cell(row=row, column=1).value is None:
                isNextMonth = False

    # convert to numpy arrays
    oil_hist = np.array(oil_hist, dtype=float)
    gas_hist = np.array(gas_hist, dtype=float)
    water_hist = np.array(water_hist, dtype=float)
    if isInj:
        waterinj_hist = np.array(waterinj_hist, dtype=float)

    # calculate GOR and Water cut
    wc = np.divide(water_hist, oil_hist + water_hist)
    gor = np.divide(gas_hist * 1000, oil_hist)

    print('Historical data for {} field has been extracted.'.format(FIELD_NAME))

    #------------------------------------------------------------------------------
    # PLOTTING
    #-------------------------------------------------------------------------------
    arial_font = fm.FontProperties(family='arial', size=8)
    # arial_font.set_family('arial')

    # 2 subplots
    fig, (ax1, ax2) = plt.subplots(2, 1, sharex=True)

    # subplot 1 (rates)
    ln1 = ax1.semilogy(months_hist, oil_hist, 'g-', label='Oil')
    ln2 = ax1.semilogy(months_hist, water_hist, 'b-', label='Water')
    if isInj:
        ln6 = ax1.semilogy(monthsinj_hist, waterinj_hist, color='royalblue', label='Water Injection')
    # ax1.yaxis.grid(which="major", linestyle='-', linewidth=1)
    # ax1.yaxis.grid(which="minor", linestyle='-', linewidth=0.5)
    # ax1.set_xlim(years_dt[0], years_dt[-1])
    ax1.xaxis.set_major_locator(years_loc)
    ax1.xaxis.set_major_formatter(yearsFmt)
    ax1.xaxis.set_minor_locator(months_loc)
    ax1.set_xlabel('Year', fontname='Arial', fontsize=8)
    ax1.set_ylabel('Liquid Rate (BPD)', fontname='Arial', fontsize=8)
    for tick in ax1.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax1.get_yticklabels():
        tick.set_fontname("Arial")

    # set figure title as ax1 title
    ax1.set_title('{} Field, Colombia History'.format(FIELD_NAME), fontname='Arial', fontsize=8)

    # secondary y-axis
    ax1_twin = ax1.twinx()
    ln3 = ax1_twin.plot(months_hist, gas_hist/1000, 'r-', label='Gas')
    ax1_twin.set_ylabel('Gas Rate (MMCFD)', color='r', fontname='Arial', fontsize=8)
    for tick in ax1_twin.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax1_twin.get_yticklabels():
        tick.set_fontname("Arial")

    # legend display (subplot 1)
    lns = ln1 + ln2 + ln3
    if isInj:
        lns = ln1 + ln2 + ln3 + ln6
    labs = [l.get_label() for l in lns]
    ax1.legend(lns, labs, loc=0, prop=arial_font)

    # subplot 2
    ln4 = ax2.plot(months_hist, gor, color='darkred',label='GOR')
    # ax1.yaxis.grid(which="major", linestyle='-', linewidth=1)
    # ax1.yaxis.grid(which="minor", linestyle='-', linewidth=0.5)
    # ax1.set_xlim(years_dt[0], years_dt[-1])
    ax2.xaxis.set_major_locator(years_loc)
    ax2.xaxis.set_major_formatter(yearsFmt)
    ax2.xaxis.set_minor_locator(months_loc)
    ax2.set_xlabel('Year', fontname='Arial', fontsize=8)
    ax2.set_ylabel('Gas Oil Ratio (SCF/STB)', fontname='Arial', fontsize=8)
    for tick in ax2.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax2.get_yticklabels():
        tick.set_fontname("Arial")

    # secondary y-axis
    ax2_twin = ax2.twinx()
    ln5 = ax2_twin.plot(months_hist, wc, color='steelblue', label='Water Cut')
    ax2_twin.set_ylabel('Water Cut', color='steelblue', fontname='Arial', fontsize=8)
    for tick in ax2_twin.get_xticklabels():
        tick.set_fontname("Arial")
    for tick in ax2_twin.get_yticklabels():
        tick.set_fontname("Arial")

    # legend display (subplot 2)
    lns2 = ln4 + ln5
    labs = [l.get_label() for l in lns2]
    ax2.legend(lns2, labs, loc=8, prop=arial_font)

    # Date labels look better
    fig.autofmt_xdate()

    fig.set_tight_layout(True)

    #plt.show()
    fig.savefig('{}_history'.format(FIELD_NAME), dpi=300)

    return None

# Check to see if this file is being executed as the "Main" python
# script instead of being used as a module by some other python script
# This allows us to use the module which ever way we want.
if __name__ == '__main__':
    CreateHistoryPlot()
